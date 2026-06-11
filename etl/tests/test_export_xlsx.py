"""Unit tests for the Layer 0 DB→xlsx review export — pure serialization +
workbook assembly only. The live-DB read (`collect`/`discover_tables`/
`fetch_table`) runs against Neon by Andy's hand (container egress to Neon is
blocked); these tests feed synthetic rows through the pure path so they need no
database, mirroring `test_validate_layer0.py`."""
from __future__ import annotations

import json
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from etl.layer0 import export_xlsx as ex


# ─── cell_value ───────────────────────────────────────────────────────


def test_cell_value_scalars_pass_through() -> None:
    assert ex.cell_value(None) is None
    assert ex.cell_value(True) is True
    assert ex.cell_value(False) is False
    assert ex.cell_value(42) == 42
    assert ex.cell_value(1.5) == 1.5
    assert ex.cell_value("Adventure Racing") == "Adventure Racing"


def test_cell_value_decimal_becomes_float() -> None:
    out = ex.cell_value(Decimal("35.5"))
    assert out == 35.5
    assert isinstance(out, float)


def test_cell_value_tz_aware_datetime_becomes_iso_string() -> None:
    # The load-bearing case: TIMESTAMPTZ → tz-aware datetime, which openpyxl
    # refuses to write. Must flatten to a string.
    dt = datetime(2026, 6, 11, 5, 1, 38, tzinfo=timezone.utc)
    out = ex.cell_value(dt)
    assert isinstance(out, str)
    assert out == dt.isoformat()


def test_cell_value_text_array_is_comma_joined() -> None:
    assert ex.cell_value(["squat", "hinge", "push"]) == "squat, hinge, push"
    assert ex.cell_value([]) == ""


def test_cell_value_jsonb_dict_is_json() -> None:
    out = ex.cell_value({"substitute_text": "band pull-apart", "is_improvised": False})
    assert json.loads(out) == {"substitute_text": "band pull-apart", "is_improvised": False}


def test_cell_value_jsonb_list_of_objects_is_json() -> None:
    payload = [{"a": 1}, {"b": 2}]
    out = ex.cell_value(payload)
    assert json.loads(out) == payload


# ─── safe_sheet_name ──────────────────────────────────────────────────


def test_safe_sheet_name_truncates_to_31() -> None:
    used: set[str] = set()
    name = ex.safe_sheet_name("a" * 50, used)
    assert len(name) == 31


def test_safe_sheet_name_replaces_reserved_chars() -> None:
    used: set[str] = set()
    name = ex.safe_sheet_name("layer0:exercises/v1", used)
    assert ":" not in name and "/" not in name


def test_safe_sheet_name_dedupes() -> None:
    used: set[str] = set()
    first = ex.safe_sheet_name("exercises", used)
    second = ex.safe_sheet_name("exercises", used)
    assert first != second
    assert first == "exercises"


# ─── build_workbook ───────────────────────────────────────────────────


def _roundtrip(wb):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf)


def test_build_workbook_one_sheet_per_table_with_header() -> None:
    tables = {
        "sports": (["id", "sport_name"], [(1, "Adventure Racing"), (2, "Skimo")]),
        "terrain_types": (["id", "canonical_name"], [(1, "Singletrack")]),
    }
    wb = _roundtrip(ex.build_workbook(tables))
    assert wb.sheetnames == ["sports", "terrain_types"]
    sports = wb["sports"]
    assert [c.value for c in sports[1]] == ["id", "sport_name"]
    assert sports["B2"].value == "Adventure Racing"
    assert sports.freeze_panes == "A2"


def test_build_workbook_serializes_complex_cells_end_to_end() -> None:
    dt = datetime(2026, 6, 11, 5, 0, 0, tzinfo=timezone.utc)
    tables = {
        "exercises": (
            ["id", "movement_patterns", "etl_run_at", "age_ramp_pct"],
            [(1, ["squat", "hinge"], dt, Decimal("0.9"))],
        ),
    }
    # Must not raise (the tz-aware datetime would crash a naive writer).
    wb = _roundtrip(ex.build_workbook(tables))
    ws = wb["exercises"]
    assert ws["B2"].value == "squat, hinge"
    assert ws["C2"].value == dt.isoformat()
    assert ws["D2"].value == 0.9


def test_build_workbook_empty_is_still_saveable() -> None:
    wb = _roundtrip(ex.build_workbook({}))
    assert wb.sheetnames == ["layer0"]
