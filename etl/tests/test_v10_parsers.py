"""Tests for the v10-specific parsers in sports_framework.

Covers `_parse_constituent_movements`, `_parse_bool`,
`_split_phase_load_notes`, `_parse_weekly_total_text`,
plus the discipline substitutes / training gaps / cross-sport
properties extractors against the in-repo v10 workbook.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from etl.layer0.extractors.sports_framework import (
    _parse_bool,
    _parse_constituent_movements,
    _parse_weekly_total_text,
    _split_phase_load_notes,
    extract_cross_sport_properties,
    extract_discipline_substitutes,
    extract_discipline_training_gaps,
    extract_phase_load_weekly_totals,
)

V10_PATH = Path(__file__).parent.parent / "sources" / "Sports_Framework_v10.xlsx"


# ---------------------------------------------------------------------------
# _parse_constituent_movements
# ---------------------------------------------------------------------------

def test_constituent_movements_happy():
    tokens, bad = _parse_constituent_movements("running;cycling")
    assert tokens == ["running", "cycling"]
    assert bad == []


def test_constituent_movements_unknown_token_surfaced():
    tokens, bad = _parse_constituent_movements("running;flying")
    assert tokens == ["running", "flying"]
    assert bad == ["flying"]


def test_constituent_movements_empty():
    assert _parse_constituent_movements(None) == (None, [])
    assert _parse_constituent_movements("") == (None, [])


def test_constituent_movements_strips_whitespace():
    tokens, bad = _parse_constituent_movements(" running ; hiking ;  ")
    assert tokens == ["running", "hiking"]
    assert bad == []


# ---------------------------------------------------------------------------
# _parse_bool
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("inp", ["TRUE", "True", "true", "T", "1", "YES", "Yes", "yes"])
def test_parse_bool_true(inp):
    assert _parse_bool(inp) is True


@pytest.mark.parametrize("inp", ["FALSE", "False", "false", "F", "0", "NO", "no"])
def test_parse_bool_false(inp):
    assert _parse_bool(inp) is False


def test_parse_bool_native_bool_passthrough():
    assert _parse_bool(True) is True
    assert _parse_bool(False) is False


def test_parse_bool_junk_returns_none():
    assert _parse_bool("maybe") is None
    assert _parse_bool(None) is None


# ---------------------------------------------------------------------------
# _split_phase_load_notes
# ---------------------------------------------------------------------------

def test_split_phase_load_notes_extracts_prescription():
    text = (
        "Standalone nav sessions only. Nav embedded within other discipline sessions "
        "is additional. Don't double-count."
    )
    presc, audit = _split_phase_load_notes(text)
    assert presc is not None
    assert presc.startswith("Standalone nav sessions only")
    assert audit is not None  # rest of sentences captured


def test_split_phase_load_notes_pure_audit_returns_none_prescription():
    text = "[AUDIT 2026-05-06] Patch confirmed."
    presc, audit = _split_phase_load_notes(text)
    assert presc is None
    assert audit is not None
    assert "AUDIT" in audit


def test_split_phase_load_notes_caps_at_120_chars():
    sentence = "x" * 200 + "."
    presc, _ = _split_phase_load_notes(sentence)
    assert presc is not None
    assert len(presc) <= 120


def test_split_phase_load_notes_empty():
    assert _split_phase_load_notes(None) == (None, None)
    assert _split_phase_load_notes("") == (None, None)


def test_split_phase_load_notes_mixed_separates():
    text = (
        "High injury-per-hour risk — quality over quantity. "
        "TAPER 12-15%: 2 short quality runs. "
        "[AUDIT 2026-05-06: Taper feasibility verified.]"
    )
    presc, audit = _split_phase_load_notes(text)
    assert presc is not None
    assert "injury-per-hour" in presc
    assert audit is not None
    assert "AUDIT" in audit


# ---------------------------------------------------------------------------
# _parse_weekly_total_text
# ---------------------------------------------------------------------------

def test_weekly_totals_happy_path():
    text = "WEEKLY TARGET HOURS: BASE: 6–9 hrs | BUILD: 7–11 hrs | PEAK: 8–12 hrs | TAPER: 5–7 hrs"
    out = _parse_weekly_total_text(text)
    assert out == {
        "Base": (6.0, 9.0),
        "Build": (7.0, 11.0),
        "Peak": (8.0, 12.0),
        "Taper": (5.0, 7.0),
    }


def test_weekly_totals_three_phases_returns_none():
    text = "BASE: 6–9 hrs | BUILD: 7–11 hrs | PEAK: 8–12 hrs"
    assert _parse_weekly_total_text(text) is None


def test_weekly_totals_non_matching_text():
    assert _parse_weekly_total_text("nothing parseable here") is None
    assert _parse_weekly_total_text(None) is None


def test_weekly_totals_single_value_phase():
    text = "BASE: ~18 hrs\nBUILD: ~22-28 hrs\nPEAK: ~28-35 hrs\nTAPER: ~12-14 hrs"
    out = _parse_weekly_total_text(text)
    assert out is not None
    assert out["Base"] == (18.0, 18.0)
    assert out["Taper"] == (12.0, 14.0)


# ---------------------------------------------------------------------------
# Workbook-backed integration tests (lightweight — read-only)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def v10_wb():
    return load_workbook(str(V10_PATH), read_only=False, data_only=True)


def test_extract_discipline_substitutes_count_91(v10_wb):
    warns: list = []
    rows = extract_discipline_substitutes(v10_wb, parse_warnings=warns)
    assert len(rows) == 91
    assert warns == []
    # Spot check one known row
    first = rows[0]
    assert first["target_id"] == "D-001"
    assert 0.0 <= first["fidelity"] <= 1.0


def test_extract_discipline_training_gaps_count_3(v10_wb):
    rows = extract_discipline_training_gaps(v10_wb)
    assert len(rows) == 3
    by_id = {r["discipline_id"]: r for r in rows}
    assert "D-018" in by_id
    assert "D-020" in by_id
    assert "D-024" in by_id
    # Swimrun → multi-substitute candidate
    assert by_id["D-018"]["multi_substitute_candidate"] is True
    assert by_id["D-018"]["gap_type"] == "no_single_substitute"
    # Alpine Descent → off-snow
    assert by_id["D-020"]["gap_type"] == "no_off_environment_substitute"


def test_extract_cross_sport_properties_filters_commentary(v10_wb):
    rows = extract_cross_sport_properties(v10_wb["Cross-Sport Properties"])
    assert len(rows) == 1
    r = rows[0]
    assert r["property_id"] == "LIT_RATIO_001"
    assert r["confidence"] == "Medium"
    assert r["source_text"] is not None
    assert r["notes"] is not None


def test_extract_phase_load_weekly_totals_emits_4_rows_per_sport(v10_wb):
    failures: list = []
    rows = extract_phase_load_weekly_totals(
        v10_wb["Phase Load Allocation"], parse_failures=failures,
    )
    # Group by sport — every parsed sport should yield exactly 4 phase rows.
    by_sport: dict[str, set[str]] = {}
    for r in rows:
        by_sport.setdefault(r["sport_name"], set()).add(r["phase"])
    for sport, phases in by_sport.items():
        assert phases == {"Base", "Build", "Peak", "Taper"}, sport


def test_substitute_with_missing_fidelity_dropped_with_warning():
    """Synthesise a workbook in memory that has a row with NULL fidelity."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Discipline Substitution Map"
    ws.append(["Target ID", "Target Name", "Substitute ID", "Substitute Name",
               "Fidelity", "Constraints", "Category"])
    ws.append(["D-001", "Trail Running", "D-003", "Hiking", 0.85, "Same musc.", "Foot"])
    ws.append(["D-002", "Road Running", "D-001", "Trail Running", None, "Bad row", "Foot"])
    warns: list = []
    rows = extract_discipline_substitutes(wb, parse_warnings=warns)
    assert len(rows) == 1
    assert rows[0]["target_id"] == "D-001"
    assert len(warns) == 1
    assert warns[0]["target_id"] == "D-002"
    assert "fidelity" in warns[0]["reason"]
