"""Tests for the sports_framework parsers against the live v14 workbook.

Covers `_parse_constituent_movements`, `_parse_bool`,
`_split_phase_load_notes`, `_parse_weekly_total_text`,
plus the discipline substitutes / training gaps / cross-sport
properties extractors against the in-repo v14 workbook.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from openpyxl import Workbook

from etl.layer0.extractors.sports_framework import (
    _parse_bool,
    _parse_constituent_movements,
    _parse_weekly_total_text,
    _split_phase_load_notes,
    extract_craft_discipline_aliases,
    extract_cross_sport_properties,
    extract_discipline_pairing_matrix,
    extract_discipline_substitutes,
    extract_discipline_training_gaps,
    extract_phase_load_weekly_totals,
)


# ---------------------------------------------------------------------------
# extract_discipline_pairing_matrix — R6 collapse dedup
# ---------------------------------------------------------------------------

def _pairing_ws_with_collapse_dup():
    """Matrix where one survivor id (D-010) appears in two header columns and
    two from-rows — simulating the R6 craft collapse (D-008a/b → D-010)."""
    wb = Workbook()
    ws = wb.active
    # Header on R10: col1 label, then destination ids (D-010 twice).
    ws.cell(row=10, column=1, value="FROM \\ TO")
    ws.cell(row=10, column=2, value="D-001")
    ws.cell(row=10, column=3, value="D-010")
    ws.cell(row=10, column=4, value="D-010")  # duplicate survivor column
    # Data R11+: D-001 row, then two D-010 from-rows (the collapse).
    ws.cell(row=11, column=1, value="D-001")
    ws.cell(row=11, column=3, value="ACC")
    ws.cell(row=11, column=4, value="ACC")    # same (D-001, D-010) again
    ws.cell(row=12, column=1, value="D-010")
    ws.cell(row=12, column=2, value="PRE")
    ws.cell(row=12, column=3, value="N/A")    # self-pair (D-010, D-010)
    ws.cell(row=13, column=1, value="D-010")  # duplicate from-row
    ws.cell(row=13, column=2, value="ACC")    # same (D-010, D-001) again
    return ws


def test_pairing_matrix_dedupes_collapse_and_skips_self_pairs():
    rows = extract_discipline_pairing_matrix(_pairing_ws_with_collapse_dup())
    pairs = [(r["discipline_id_a"], r["discipline_id_b"]) for r in rows]
    # No duplicate (a, b) pairs survive the collapse.
    assert len(pairs) == len(set(pairs)), pairs
    # No self-pair from the collapsed diagonal.
    assert ("D-010", "D-010") not in pairs
    # First-seen-wins: (D-001, D-010) once = ACCEPTABLE; (D-010, D-001) = PREFERRED.
    by_pair = {(r["discipline_id_a"], r["discipline_id_b"]): r["pairing_rating"] for r in rows}
    assert by_pair[("D-001", "D-010")] == "ACCEPTABLE"
    assert by_pair[("D-010", "D-001")] == "PREFERRED"
    assert pairs.count(("D-001", "D-010")) == 1

V14_PATH = Path(__file__).parent.parent / "sources" / "Sports_Framework_v14.xlsx"


# ---------------------------------------------------------------------------
# _parse_constituent_movements
# ---------------------------------------------------------------------------

def test_constituent_movements_happy():
    tokens, bad = _parse_constituent_movements("running;cycling")
    assert tokens == ["running", "cycling"]
    assert bad == []


def test_constituent_movements_unknown_token_surfaced():
    tokens, bad = _parse_constituent_movements("running;flying")
    assert tokens == ["running", "flying"]
    assert bad == ["flying"]


def test_constituent_movements_empty():
    assert _parse_constituent_movements(None) == (None, [])
    assert _parse_constituent_movements("") == (None, [])


def test_constituent_movements_strips_whitespace():
    tokens, bad = _parse_constituent_movements(" running ; hiking ;  ")
    assert tokens == ["running", "hiking"]
    assert bad == []


# ---------------------------------------------------------------------------
# _parse_bool
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("inp", ["TRUE", "True", "true", "T", "1", "YES", "Yes", "yes"])
def test_parse_bool_true(inp):
    assert _parse_bool(inp) is True


@pytest.mark.parametrize("inp", ["FALSE", "False", "false", "F", "0", "NO", "no"])
def test_parse_bool_false(inp):
    assert _parse_bool(inp) is False


def test_parse_bool_native_bool_passthrough():
    assert _parse_bool(True) is True
    assert _parse_bool(False) is False


def test_parse_bool_junk_returns_none():
    assert _parse_bool("maybe") is None
    assert _parse_bool(None) is None


# ---------------------------------------------------------------------------
# _split_phase_load_notes
# ---------------------------------------------------------------------------

def test_split_phase_load_notes_extracts_prescription():
    text = (
        "Standalone nav sessions only. Nav embedded within other discipline sessions "
        "is additional. Don't double-count."
    )
    presc, audit = _split_phase_load_notes(text)
    assert presc is not None
    assert presc.startswith("Standalone nav sessions only")
    assert audit is not None  # rest of sentences captured


def test_split_phase_load_notes_pure_audit_returns_none_prescription():
    text = "[AUDIT 2026-05-06] Patch confirmed."
    presc, audit = _split_phase_load_notes(text)
    assert presc is None
    assert audit is not None
    assert "AUDIT" in audit


def test_split_phase_load_notes_caps_at_120_chars():
    sentence = "x" * 200 + "."
    presc, _ = _split_phase_load_notes(sentence)
    assert presc is not None
    assert len(presc) <= 120


def test_split_phase_load_notes_empty():
    assert _split_phase_load_notes(None) == (None, None)
    assert _split_phase_load_notes("") == (None, None)


def test_split_phase_load_notes_mixed_separates():
    text = (
        "High injury-per-hour risk — quality over quantity. "
        "TAPER 12-15%: 2 short quality runs. "
        "[AUDIT 2026-05-06: Taper feasibility verified.]"
    )
    presc, audit = _split_phase_load_notes(text)
    assert presc is not None
    assert "injury-per-hour" in presc
    assert audit is not None
    assert "AUDIT" in audit


# ---------------------------------------------------------------------------
# _parse_weekly_total_text
# ---------------------------------------------------------------------------

def test_weekly_totals_happy_path():
    text = "WEEKLY TARGET HOURS: BASE: 6–9 hrs | BUILD: 7–11 hrs | PEAK: 8–12 hrs | TAPER: 5–7 hrs"
    out = _parse_weekly_total_text(text)
    assert out == {
        "Base": (6.0, 9.0, "hrs"),
        "Build": (7.0, 11.0, "hrs"),
        "Peak": (8.0, 12.0, "hrs"),
        "Taper": (5.0, 7.0, "hrs"),
    }


def test_weekly_totals_three_phases_returns_none():
    text = "BASE: 6–9 hrs | BUILD: 7–11 hrs | PEAK: 8–12 hrs"
    assert _parse_weekly_total_text(text) is None


def test_weekly_totals_non_matching_text():
    assert _parse_weekly_total_text("nothing parseable here") is None
    assert _parse_weekly_total_text(None) is None


def test_weekly_totals_single_value_phase():
    text = "BASE: ~18 hrs\nBUILD: ~22-28 hrs\nPEAK: ~28-35 hrs\nTAPER: ~12-14 hrs"
    out = _parse_weekly_total_text(text)
    assert out is not None
    assert out["Base"] == (18.0, 18.0, "hrs")
    assert out["Taper"] == (12.0, 14.0, "hrs")


def test_weekly_totals_km_volume():
    # Open Water Marathon Swimming (10km / Olympic). Source row R192.
    text = (
        "WEEKLY TARGET VOLUME (km/wk): BASE: 30–45 km | BUILD: 40–55 km | "
        "PEAK: 45–60 km | TAPER: 20–30 km  Volume measured in km"
    )
    out = _parse_weekly_total_text(text)
    assert out == {
        "Base": (30.0, 45.0, "km"),
        "Build": (40.0, 55.0, "km"),
        "Peak": (45.0, 60.0, "km"),
        "Taper": (20.0, 30.0, "km"),
    }


def test_weekly_totals_multi_subformat_aggregates_envelope():
    # Swimrun (R65) — three sub-formats per phase, aggregated to min/max
    # envelope across them. Parenthesized km distances must NOT be
    # mistaken for target values.
    text = (
        "WEEKLY TARGET HOURS: "
        "BASE: Sprint (10–25km): 4–6 hrs World Series (25–40km): 8–12 hrs ÖTILLÖ (75km): 12–16 hrs "
        "BUILD: Sprint (10–25km): 5–7 hrs World Series (25–40km): 10–14 hrs ÖTILLÖ (75km): 14–18 hrs "
        "PEAK: Sprint (10–25km): 6–8 hrs World Series (25–40km): 12–16 hrs ÖTILLÖ (75km): 16–20 hrs "
        "TAPER: Sprint (10–25km): 3–5 hrs World Series (25–40km): 6–9 hrs ÖTILLÖ (75km): 8–12 hrs"
    )
    out = _parse_weekly_total_text(text)
    assert out == {
        "Base": (4.0, 16.0, "hrs"),
        "Build": (5.0, 18.0, "hrs"),
        "Peak": (6.0, 20.0, "hrs"),
        "Taper": (3.0, 12.0, "hrs"),
    }


def test_weekly_totals_offroad_multisport_subformats():
    # Off-Road / Adventure Multisport (R167). Three sub-formats per phase
    # without parenthesized labels — purely whitespace-delimited.
    text = (
        "WEEKLY TARGET HOURS: "
        "BASE: XTERRA: 8–12 hrs Quadrathlon: 9–13 hrs Free-format: 8–14 hrs "
        "BUILD: XTERRA: 10–15 hrs Quadrathlon: 11–16 hrs Free-format: 10–18 hrs "
        "PEAK: XTERRA: 12–18 hrs Quadrathlon: 13–18 hrs Free-format: 12–20 hrs "
        "TAPER: XTERRA: 5–8 hrs Quadrathlon: 6–9 hrs Free-format: 5–10 hrs"
    )
    out = _parse_weekly_total_text(text)
    assert out is not None
    assert out["Base"] == (8.0, 14.0, "hrs")
    assert out["Taper"] == (5.0, 10.0, "hrs")


def test_weekly_totals_mixed_units_within_phase_rejects_phase():
    # If a phase mixes hrs and km, treat as ambiguous and skip — the row
    # then fails the 4-phase quorum.
    text = "BASE: 6–9 hrs and 10–20 km | BUILD: 7–11 hrs | PEAK: 8–12 hrs | TAPER: 5–7 hrs"
    out = _parse_weekly_total_text(text)
    assert out is None


# ---------------------------------------------------------------------------
# Workbook-backed integration tests (lightweight — read-only)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def v14_wb():
    return load_workbook(str(V14_PATH), read_only=False, data_only=True)


def test_extract_discipline_substitutes_count_89(v14_wb):
    # 89 = 91 source rows − 2 dropped by the R6 collapse dedup (one duplicate
    # (D-010, D-011, 'Canoeing') key + one self-substitute the merge created).
    warns: list = []
    rows = extract_discipline_substitutes(v14_wb, parse_warnings=warns)
    assert len(rows) == 89
    assert warns == []
    # No duplicate UNIQUE keys + no self-substitutes survive the dedup.
    keys = [(r["target_id"], r["substitute_id"], r["substitute_name"]) for r in rows]
    assert len(keys) == len(set(keys))
    assert not any(r["target_id"] == r["substitute_id"] for r in rows)
    # Spot check one known row
    first = rows[0]
    assert first["target_id"] == "D-001"
    assert 0.0 <= first["fidelity"] <= 1.0


def test_extract_discipline_training_gaps_count_3(v14_wb):
    rows = extract_discipline_training_gaps(v14_wb)
    assert len(rows) == 3
    by_id = {r["discipline_id"]: r for r in rows}
    assert "D-020" in by_id
    assert "D-022" in by_id
    assert "D-025" in by_id
    # Swimrun → multi-substitute candidate
    assert by_id["D-020"]["multi_substitute_candidate"] is True
    assert by_id["D-020"]["gap_type"] == "no_single_substitute"
    # Alpine Descent → off-snow
    assert by_id["D-022"]["gap_type"] == "no_off_environment_substitute"


def test_extract_cross_sport_properties_filters_commentary(v14_wb):
    rows = extract_cross_sport_properties(v14_wb["Cross-Sport Properties"])
    assert len(rows) == 1
    r = rows[0]
    assert r["property_id"] == "LIT_RATIO_001"
    assert r["confidence"] == "Medium"
    assert r["source_text"] is not None
    assert r["notes"] is not None


def test_extract_phase_load_weekly_totals_emits_4_rows_per_sport(v14_wb):
    failures: list = []
    rows = extract_phase_load_weekly_totals(
        v14_wb["Phase Load Allocation"], parse_failures=failures,
    )
    # Group by sport — every parsed sport should yield exactly 4 phase rows.
    by_sport: dict[str, set[str]] = {}
    for r in rows:
        by_sport.setdefault(r["sport_name"], set()).add(r["phase"])
    for sport, phases in by_sport.items():
        assert phases == {"Base", "Build", "Peak", "Taper"}, sport


def test_substitute_with_missing_fidelity_dropped_with_warning():
    """Synthesise a workbook in memory that has a row with NULL fidelity."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Discipline Substitution Map"
    ws.append(["Target ID", "Target Name", "Substitute ID", "Substitute Name",
               "Fidelity", "Constraints", "Category"])
    ws.append(["D-001", "Trail Running", "D-003", "Hiking", 0.85, "Same musc.", "Foot"])
    ws.append(["D-002", "Road Running", "D-001", "Trail Running", None, "Bad row", "Foot"])
    warns: list = []
    rows = extract_discipline_substitutes(wb, parse_warnings=warns)
    assert len(rows) == 1
    assert rows[0]["target_id"] == "D-001"
    assert len(warns) == 1
    assert warns[0]["target_id"] == "D-002"
    assert "fidelity" in warns[0]["reason"]


# ---------------------------------------------------------------------------
# X1b.3b — extract_craft_discipline_aliases (Craft Discipline Aliases sheet)
# ---------------------------------------------------------------------------

def test_extract_craft_discipline_aliases_seed(v14_wb):
    rows = extract_craft_discipline_aliases(v14_wb)
    by_craft: dict[str, set[str]] = {}
    for r in rows:
        assert set(r) == {"craft_name", "discipline_id", "group_kind"}
        by_craft.setdefault(r["craft_name"], set()).add(r["discipline_id"])
    # Many-to-many: gravel bike trains road + gravel + XC; trainer = all bikes.
    assert by_craft["kayak"] == {"D-010"}
    assert by_craft["gravel_bike"] == {"D-006", "D-030", "D-031"}
    assert by_craft["mountain_bike"] == {"D-008", "D-031"}
    assert by_craft["cycling_trainer"] == {"D-006", "D-007", "D-008", "D-030", "D-031"}
    assert len(rows) == 14


def test_extract_craft_discipline_aliases_validates_group_kind():
    from openpyxl import Workbook

    from etl.layer0.extractors.sports_framework import _VALID_GROUP_KINDS

    wb = Workbook()
    wb.active.title = "Craft Discipline Aliases"
    ws = wb["Craft Discipline Aliases"]
    ws.append(["Craft Slug", "Discipline ID", "Group Kind"])
    ws.append(["road_bike", "D-006", "not_a_kind"])
    assert "not_a_kind" not in _VALID_GROUP_KINDS
    with pytest.raises(ValueError):
        extract_craft_discipline_aliases(wb)


def test_extract_craft_discipline_aliases_absent_sheet_back_compat():
    from openpyxl import Workbook

    wb = Workbook()  # no "Craft Discipline Aliases" sheet
    assert extract_craft_discipline_aliases(wb) == []
