#!/usr/bin/env python3
"""R6 discipline-ID renumber + two collapses — workbook transform.

Produces `Sports_Framework_v11.xlsx` from `Sports_Framework_v10.xlsx` by:
  1. Renumbering every discipline-id token (`D-\\d{3}[a-z]?`) in every
     string cell of every sheet, per the authoritative R6 map (see
     `aidstation-sources/Discipline_ID_Renumber_R6_Design_v1.md` §2).
  2. Collapsing the Discipline Library sheet's two merged pairs to one
     survivor row each (kayak flat/whitewater -> D-010 "Kayaking";
     mountain-running uphill/downhill -> D-024 "Mountain Running").

What this script intentionally does NOT do (owed to the Neon re-extract,
which needs the live ETL loop to validate — see design §7):
  - Deduping the duplicate D-010 / D-024 headers + rows the collapse
    creates in the Discipline Pairing Matrix, Substitution Map,
    Training Gaps, and Cross-Sport Properties sheets. The ETL's per-table
    UNIQUE constraints + first-seen-wins dedup handle (sport, discipline_id)
    collisions on load (see `extract_sport_discipline_map`); the matrix /
    substitution collapse-merges should be reviewed against the
    pairing-count guard after the ETL run.

Usage:  python3 etl/sources/migrate_discipline_ids_R6.py
"""
from __future__ import annotations

import os
import re

import openpyxl

MAP = {
    "D-004b": "D-005", "D-005": "D-006", "D-005a": "D-007", "D-006": "D-008",
    "D-007": "D-009",
    "D-008a": "D-010", "D-008b": "D-010",            # COLLAPSE -> Kayaking
    "D-009": "D-011", "D-010": "D-012", "D-011": "D-013", "D-012": "D-014",
    "D-013": "D-015", "D-014": "D-016", "D-015": "D-017", "D-016": "D-018",
    "D-017": "D-019", "D-018": "D-020", "D-019": "D-021", "D-020": "D-022",
    "D-021": "D-023",
    "D-022": "D-024", "D-023": "D-024",              # COLLAPSE -> Mountain Running
    "D-024": "D-025", "D-025": "D-026", "D-026": "D-027",
    # D-001..D-004, D-028, D-029 unchanged (identity, omitted).
}
SURVIVOR_NAMES = {"D-010": "Kayaking", "D-024": "Mountain Running"}
TOKEN = re.compile(r"D-\d{3}[a-z]?")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
SRC = os.path.join(HERE, "Sports_Framework_v10.xlsx")
OUTS = [
    os.path.join(HERE, "Sports_Framework_v11.xlsx"),
    os.path.join(ROOT, "aidstation-sources", "data", "Sports_Framework_v11.xlsx"),
]


def _renumber(text: str) -> str:
    return TOKEN.sub(lambda m: MAP.get(m.group(0), m.group(0)), text)


def main() -> None:
    wb = openpyxl.load_workbook(SRC)
    cell_rewrites = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and TOKEN.search(cell.value):
                    new = _renumber(cell.value)
                    if new != cell.value:
                        cell.value = new
                        cell_rewrites += 1

    # Collapse the Discipline Library rows (the one sheet the ETL does not
    # dedupe — extract_disciplines appends every D- row).
    lib = wb["Discipline Library"]
    seen: set[str] = set()
    drop_rows: list[int] = []
    for r in range(2, lib.max_row + 1):
        did = lib.cell(row=r, column=1).value
        if not (isinstance(did, str) and did.startswith("D-")):
            continue
        if did in seen:
            drop_rows.append(r)
            continue
        seen.add(did)
        if did in SURVIVOR_NAMES:
            lib.cell(row=r, column=2).value = SURVIVOR_NAMES[did]
    for r in reversed(drop_rows):   # delete bottom-up to keep indices valid
        lib.delete_rows(r, 1)

    for out in OUTS:
        wb.save(out)
    print(f"cell rewrites: {cell_rewrites}; collapsed rows dropped: {len(drop_rows)}; "
          f"disciplines: {len(seen)}")
    print("wrote:", ", ".join(os.path.relpath(o, ROOT) for o in OUTS))


if __name__ == "__main__":
    main()
