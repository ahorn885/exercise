"""Layer 0 ETL — extractor for AR_Exercise_Database_v17.xlsx (source 0B).

Reads Exercise Master + Sport-Exercise Map (header on R2 in both per spec).
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from etl.layer0.vocabulary_transforms import (
    transform_body_part_string,
    transform_equipment_string,
)

# `EX### — Name` parser. The em-dash (—) is the canonical separator;
# fall back to "-" / "–" to be resilient.
_EX_REF = re.compile(r"^\s*(EX\d+)\s*[—–\-]\s*(.+?)\s*$")


def _t(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _i(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None


def _split_comma(text: str | None) -> list[str]:
    if not text:
        return []
    return [p.strip() for p in str(text).split(",") if p.strip()]


def parse_exercise_ref(text: str | None) -> dict[str, str | None]:
    """Parse `EX### — Name` → {exercise_id, exercise_name}.

    Returns {"exercise_id": None, "exercise_name": None} on no match.
    Spec §4.10: used for progression / regression / physical_proxies entries.
    """
    if not text:
        return {"exercise_id": None, "exercise_name": None}
    m = _EX_REF.match(str(text))
    if not m:
        return {"exercise_id": None, "exercise_name": None}
    return {"exercise_id": m.group(1), "exercise_name": m.group(2).strip()}


def parse_physical_proxies(text: str | None) -> list[dict[str, str | None]]:
    """Spec §4.10: split on `;`, then parse each `EX### — Name`."""
    if not text:
        return []
    out: list[dict[str, str | None]] = []
    for chunk in str(text).split(";"):
        ref = parse_exercise_ref(chunk.strip())
        if ref["exercise_id"]:
            out.append(ref)
    return out


def parse_equipment_substitutes(text: str | None) -> dict[str, list[str]]:
    """Spec §4.10: split on `;`. Each entry is either a standard substitute
    or, with 🏠 prefix, an improvised option (the prefix is stripped)."""
    if not text:
        return {"standard": [], "improvised": []}
    standard: list[str] = []
    improvised: list[str] = []
    for chunk in str(text).split(";"):
        item = chunk.strip()
        if not item:
            continue
        if item.startswith("🏠"):
            stripped = item.lstrip("🏠").strip()
            if stripped:
                improvised.append(stripped)
        else:
            standard.append(item)
    return {"standard": standard, "improvised": improvised}


def open_workbook(path: str | Path):
    return load_workbook(str(path), read_only=False, data_only=True)


def extract_exercises(ws: Worksheet) -> list[dict[str, Any]]:
    """Exercise Master sheet. Header on R2; data R3+.

    Applies vocabulary_transforms.transform_equipment_string() to col 7
    (Equipment) before storing — per spec §3.1, §4.10.
    """
    rows: list[dict[str, Any]] = []
    for r in range(3, ws.max_row + 1):
        ex_id = _t(ws.cell(row=r, column=1).value)
        if not ex_id or not ex_id.startswith("EX"):
            continue

        progression = parse_exercise_ref(_t(ws.cell(row=r, column=14).value))
        regression = parse_exercise_ref(_t(ws.cell(row=r, column=15).value))
        equipment_raw = _t(ws.cell(row=r, column=7).value)
        equipment_canonical = transform_equipment_string(equipment_raw)

        rows.append({
            "exercise_id": ex_id,
            "exercise_name": _t(ws.cell(row=r, column=2).value) or "",
            "exercise_type": _t(ws.cell(row=r, column=3).value) or "",
            "movement_patterns": _split_comma(ws.cell(row=r, column=4).value),
            "primary_muscles": _split_comma(ws.cell(row=r, column=5).value),
            "secondary_muscles": _split_comma(ws.cell(row=r, column=6).value),
            "equipment_required": equipment_canonical,
            # col 7 (Novelty) excluded entirely per spec §4.10
            "injury_flags_text": _t(ws.cell(row=r, column=9).value),
            # col 9 = "Notes / Coaching Cues" → coaching_cues field
            "coaching_cues": _t(ws.cell(row=r, column=10).value),
            "equipment_substitutes": parse_equipment_substitutes(
                _t(ws.cell(row=r, column=11).value)
            ),
            "physical_proxies": parse_physical_proxies(_t(ws.cell(row=r, column=12).value)),
            "contraindicated_parts": transform_body_part_string(
                _t(ws.cell(row=r, column=13).value)
            ),
            "progression_exercise_id": progression["exercise_id"],
            "progression_exercise_name": progression["exercise_name"],
            "regression_exercise_id": regression["exercise_id"],
            "regression_exercise_name": regression["exercise_name"],
            "sport_count": _i(ws.cell(row=r, column=16).value),
        })
    return rows


def extract_sport_exercise_map(
    ws: Worksheet,
    dropped_dupes: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Sport-Exercise Map sheet. Header on R2; data R3+.

    Dedupes by `(exercise_id, sport_name)` to satisfy the spec's UNIQUE
    constraint. First-seen wins. As of v17 the source has 3 such pairs
    (EX163/Canoeing, EX023/Fencing, EX207/XC Skiing) where the same
    exercise is listed twice with different relevance notes — likely
    accidental rephrasings during DB curation.
    """
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for r in range(3, ws.max_row + 1):
        ex_id = _t(ws.cell(row=r, column=1).value)
        if not ex_id or not ex_id.startswith("EX"):
            continue
        sport = _t(ws.cell(row=r, column=4).value) or ""
        row = {
            "exercise_id": ex_id,
            "exercise_name": _t(ws.cell(row=r, column=2).value) or "",
            "exercise_type": _t(ws.cell(row=r, column=3).value) or "",
            "sport_name": sport,
            "sport_relevance_note": _t(ws.cell(row=r, column=5).value) or "",
            "priority": _t(ws.cell(row=r, column=6).value) or "",
        }
        key = (ex_id, sport)
        if key in seen:
            if dropped_dupes is not None:
                dropped_dupes.append({
                    "row_number": r,
                    "exercise_id": ex_id,
                    "sport_name": sport,
                    "exercise_name": row["exercise_name"],
                    "priority": row["priority"],
                })
            continue
        seen.add(key)
        rows.append(row)
    return rows
