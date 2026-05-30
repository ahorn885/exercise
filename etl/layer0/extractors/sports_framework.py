"""Layer 0 ETL — extractor for Sports_Framework_v11.xlsx (source 0A).

One function per sheet. Each returns a list of dicts ready for INSERT.
Parsing rules per spec §4.2–§4.15 (Layer0_ETL_Spec_v3.md).
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Regex helpers
# ---------------------------------------------------------------------------

# en-dash (–) and hyphen-minus (-) both observed in source data
_DASH = r"[–\-]"

_PACK_WEIGHT_RE = re.compile(rf"(\d+)(?:\s*{_DASH}\s*(\d+))?\s*lb", re.IGNORECASE)
_WEEKS_RE = re.compile(rf"(\d+)(?:\s*{_DASH}\s*(\d+))?\s*weeks?", re.IGNORECASE)
_PCT_RE = re.compile(rf"(\d+(?:\.\d+)?)(?:\s*{_DASH}\s*(\d+(?:\.\d+)?))?\s*%")

# v10 — Sports Index col 13 enum tokens (split on `;`)
ENUM_MOVEMENTS = {
    "running", "cycling", "swimming", "paddling", "skiing",
    "climbing", "hiking", "navigation", "other_skill",
}
ENUM_ENDURANCE = {"Pure endurance", "Mixed", "Technical-dominant"}
ENUM_FORMAT = {"Individual", "Team", "Both"}
ENUM_DEFAULT_INCLUSION = {"included", "excluded", "prompt_required"}

# v10 — Phase Load Allocation Notes split heuristic
_AUDIT_PREFIXES = (
    "[", "Source:", "Audit:", "*CONDITIONAL", "PENDING",
    "[AUDIT", "*Conditional", "[TAPER feasibility patch",
)

# v10 — WEEKLY TOTAL TARGET parser
#
# Two notation families appear in the source spreadsheet:
#   1. Time-based:  "BASE: 6–9 hrs | BUILD: 7–11 hrs | ..."
#                   "BASE: ~18 hrs" (single value → low=high)
#   2. Volume-based (open-water swim only):
#                   "BASE: 30–45 km | BUILD: 40–55 km | ..."
#   3. Multi sub-format time:
#                   "BASE: Sprint (10–25km): 4–6 hrs World Series: 8–12 hrs BUILD: ..."
#                   Aggregated to one (min_low, max_high) range per phase.
#
# `_PHASE_HEAD_RE` locates phase section starts; `_RANGE_RE` finds the
# numeric ranges inside each section. Parenthesized sub-format labels are
# stripped before range matching so `(10–25km)` doesn't get mistaken for a
# target value.
_PHASE_HEAD_RE = re.compile(r"\b(?P<phase>BASE|BUILD|PEAK|TAPER)\s*:", re.IGNORECASE)
_RANGE_RE = re.compile(
    rf"~?\s*(?P<low>\d+(?:\.\d+)?)"
    rf"(?:\s*{_DASH}\s*(?P<high>\d+(?:\.\d+)?))?"
    rf"\s*(?P<unit>hrs?|km)\b",
    re.IGNORECASE,
)
_PAREN_RE = re.compile(r"\([^()]*\)")
_PHASE_CANON = {"base": "Base", "build": "Build", "peak": "Peak", "taper": "Taper"}

# v10 — Cross-Sport Properties row filter
_PROPERTY_ID_RE = re.compile(r"^[A-Z]+_[A-Z]+_\d{3}$")

# Age band header markers — used to slice the age_adjusted_ramp text
# into per-band sections. The text format is consistently
#   "40–44: ...\n45–54: ...\n55+: ..."
# but some disciplines use just "40-44:" or "Standard ramp" globally.
_AGE_BANDS = (
    ("40_44", re.compile(rf"40\s*{_DASH}\s*44\s*:")),
    ("45_54", re.compile(rf"45\s*{_DASH}\s*54\s*:")),
    ("55_plus", re.compile(r"55\s*\+\s*:")),
)


def _parse_pack_weight(text: str | None) -> tuple[float | None, float | None]:
    """Extract (low, high) lb values from pack carry notes."""
    if not text:
        return None, None
    m = _PACK_WEIGHT_RE.search(text)
    if not m:
        return None, None
    low = float(m.group(1))
    high = float(m.group(2)) if m.group(2) else low
    return low, high


def _parse_weeks(text: str | None) -> tuple[int | None, int | None]:
    if not text:
        return None, None
    m = _WEEKS_RE.search(text)
    if not m:
        return None, None
    low = int(m.group(1))
    high = int(m.group(2)) if m.group(2) else low
    return low, high


def _parse_race_time_pct(text: str | None) -> tuple[float | None, float | None]:
    if not text:
        return None, None
    m = _PCT_RE.search(text)
    if not m:
        return None, None
    low = float(m.group(1))
    high = float(m.group(2)) if m.group(2) else low
    return low, high


def _parse_age_ramps(text: str | None) -> dict[str, float | None]:
    """Slice the age-adjusted ramp text into 3 bands; extract the first %
    value from each band's slice. Returns NULLs for bands that say
    "Standard ramp" or otherwise have no % token."""
    out = {"40_44": None, "45_54": None, "55_plus": None}
    if not text:
        return out

    # Find each band header start; use them as slice boundaries.
    positions: list[tuple[str, int]] = []
    for key, pat in _AGE_BANDS:
        m = pat.search(text)
        if m:
            positions.append((key, m.end()))

    if not positions:
        return out
    positions.sort(key=lambda x: x[1])
    for i, (key, start) in enumerate(positions):
        end = positions[i + 1][1] if i + 1 < len(positions) else len(text)
        slice_text = text[start:end]
        m = _PCT_RE.search(slice_text)
        if m:
            out[key] = float(m.group(1))
    return out


def _split_dot_list(text: str | None) -> list[str]:
    """Split a ` · `-delimited list, trim entries, drop empties."""
    if not text:
        return []
    return [p.strip() for p in str(text).split(" · ") if p.strip()]


def _parse_recovery_modalities(text: str | None) -> list[str]:
    """Col 11 is a numbered list:
        '1. Sleep (8+ hrs)\n2. CWI lower body (11–15°C, 10–15 min)\n3. ...'
    Extract each item after the digit-and-period prefix.
    """
    if not text:
        return []
    items: list[str] = []
    for line in str(text).splitlines():
        m = re.match(r"^\s*\d+[.)]\s*(.+?)\s*$", line)
        if m:
            items.append(m.group(1))
    return items


def _parse_yes_no(text: str | None) -> tuple[bool, str]:
    """Parse a 'YES/NO — explanation' cell.

    Returns (flag_bool, full_text).
    """
    if text is None:
        return False, ""
    s = str(text).strip()
    if not s:
        return False, ""
    # First whitespace-delimited token, stripping a trailing dash separator
    head = re.split(r"[\s—–\-]", s, maxsplit=1)[0].strip().upper()
    flag = head == "YES"
    return flag, s


def _i(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None


def _f(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _t(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip().replace("\n", " ")
    return s or None


def _t_raw(value: Any) -> str | None:
    """Like `_t` but preserves embedded newlines (used for raw_notes)."""
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _parse_constituent_movements(raw: Any) -> tuple[list[str] | None, list[str]]:
    """Parse the Sports Index `Constituent Movements` cell.

    Returns (tokens, unknowns). `unknowns` is a list of tokens that are not
    in `ENUM_MOVEMENTS`. The caller decides whether to surface them.
    """
    if not raw:
        return None, []
    tokens = [t.strip() for t in str(raw).split(";") if t.strip()]
    if not tokens:
        return None, []
    bad = [t for t in tokens if t not in ENUM_MOVEMENTS]
    return tokens, bad


def _parse_bool(raw: Any) -> bool | None:
    if raw is None:
        return None
    if isinstance(raw, bool):
        return raw
    s = str(raw).strip().upper()
    if s in ("TRUE", "T", "1", "YES"):
        return True
    if s in ("FALSE", "F", "0", "NO"):
        return False
    return None


def _split_phase_load_notes(notes: Any) -> tuple[str | None, str | None]:
    """Split Notes / Conditions cell into (prescription_note, audit_log).

    `prescription_note` is the first sentence/clause that doesn't begin with
    an audit prefix, trimmed and capped at ~120 chars. `audit_log` is the
    remaining chunks joined with ` | `. Returns `(None, raw)` on parse failure.
    """
    if not notes:
        return None, None
    text = str(notes).strip()
    try:
        parts = re.split(r"(?<=[.;])\s+|\n+", text)
    except re.error:
        return None, text
    prescription: str | None = None
    audit_chunks: list[str] = []
    for p in parts:
        p_clean = p.strip()
        if not p_clean:
            continue
        is_audit = any(p_clean.startswith(prefix) for prefix in _AUDIT_PREFIXES)
        if not is_audit and prescription is None:
            prescription = p_clean[:120].rstrip()
        else:
            audit_chunks.append(p_clean)
    audit_log = " | ".join(audit_chunks) if audit_chunks else None
    return prescription, audit_log


def _parse_weekly_total_text(
    text: Any,
) -> dict[str, tuple[float, float, str]] | None:
    """Returns {Base: (low, high, unit), Build, Peak, Taper} or None on bad parse.

    Handles three notation families seen in the source spreadsheet:

      1. Direct time:   `BASE: 6–9 hrs | BUILD: 7–11 hrs | ...`
                        `BASE: ~18 hrs` (single value → low=high)
      2. Direct volume: `BASE: 30–45 km | BUILD: 40–55 km | ...` (open-water
                        marathon swimming is measured in km/wk, not hrs)
      3. Multi sub-format time: `BASE: Sprint (10–25km): 4–6 hrs World Series
                        (25–40km): 8–12 hrs BUILD: ...` — each phase carries
                        multiple sub-format ranges; aggregated to one envelope
                        `(min(lows), max(highs))` per phase.

    Per-phase unit must be consistent (all hrs or all km); mixing causes the
    phase to be rejected. All four phases must be populated for the row to
    be considered valid.
    """
    if not text:
        return None
    text_str = str(text)
    heads = list(_PHASE_HEAD_RE.finditer(text_str))
    if not heads:
        return None

    out: dict[str, tuple[float, float, str]] = {}
    for i, m in enumerate(heads):
        phase = _PHASE_CANON.get(m.group("phase").lower())
        if not phase or phase in out:
            continue
        section_start = m.end()
        section_end = heads[i + 1].start() if i + 1 < len(heads) else len(text_str)
        section = text_str[section_start:section_end]
        # Strip parenthesized sub-format labels like "(10–25km)" so their
        # numbers don't get mistaken for targets.
        section = _PAREN_RE.sub(" ", section)

        lows: list[float] = []
        highs: list[float] = []
        units: set[str] = set()
        for rm in _RANGE_RE.finditer(section):
            low = float(rm.group("low"))
            high = float(rm.group("high")) if rm.group("high") else low
            unit = rm.group("unit").lower()
            unit = "hrs" if unit.startswith("hr") else unit
            lows.append(low)
            highs.append(high)
            units.add(unit)

        if not lows or len(units) != 1:
            continue
        out[phase] = (min(lows), max(highs), next(iter(units)))

    if len(out) < 4:
        return None
    return out


# ---------------------------------------------------------------------------
# Sheet extractors
# ---------------------------------------------------------------------------

def open_workbook(path: str | Path):
    return load_workbook(str(path), read_only=False, data_only=True)


def extract_sports(
    ws: Worksheet,
    movement_warnings: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Sheet 1 — Sports Index. Header on R1, data R2+.

    v10 adds four classification columns at cols 13/14/15/16:
    Constituent Movements, Endurance Profile, Participation Format,
    Multi-Discipline. Unknown enum tokens are surfaced via
    `movement_warnings` if provided (informational, never fails the ETL).
    """
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        sport = _t(ws.cell(row=r, column=1).value)
        if not sport:
            continue
        nav_text = ws.cell(row=r, column=6).value
        sleep_text = ws.cell(row=r, column=7).value
        pack_text = ws.cell(row=r, column=8).value
        trans_text = ws.cell(row=r, column=9).value

        flag_nav, nav_full = _parse_yes_no(nav_text)
        flag_sleep, sleep_full = _parse_yes_no(sleep_text)
        flag_pack, pack_full = _parse_yes_no(pack_text)
        flag_trans, trans_full = _parse_yes_no(trans_text)
        pack_low, pack_high = _parse_pack_weight(pack_full) if flag_pack else (None, None)

        movements, bad_movements = _parse_constituent_movements(
            ws.cell(row=r, column=13).value
        )
        endurance = _t(ws.cell(row=r, column=14).value)
        participation = _t(ws.cell(row=r, column=15).value)
        multi_disc = _parse_bool(ws.cell(row=r, column=16).value)

        warns: list[str] = []
        if bad_movements:
            warns.append(f"unknown movement tokens: {bad_movements}")
        if endurance and endurance not in ENUM_ENDURANCE:
            warns.append(f"unknown endurance_profile {endurance!r}")
        if participation and participation not in ENUM_FORMAT:
            warns.append(f"unknown participation_format {participation!r}")
        if multi_disc is not None and movements is not None:
            derived = len(movements) > 1
            if derived != multi_disc:
                warns.append(
                    f"multi_discipline={multi_disc} disagrees with derived "
                    f"len(constituent_movements)>1={derived}"
                )
        if warns and movement_warnings is not None:
            movement_warnings.append({
                "sport_name": sport,
                "row_number": r,
                "warnings": warns,
            })

        rows.append({
            "sport_name": sport,
            "typical_duration_range": _t(ws.cell(row=r, column=4).value),
            "team_vs_solo": _t(ws.cell(row=r, column=5).value),
            "flag_navigation": flag_nav,
            "navigation_notes": nav_full or None,
            "flag_sleep_deprivation": flag_sleep,
            "sleep_deprivation_notes": sleep_full or None,
            "flag_pack_carry": flag_pack,
            "pack_carry_notes": pack_full or None,
            "pack_weight_lbs_low": pack_low,
            "pack_weight_lbs_high": pack_high,
            "flag_transition_training": flag_trans,
            "transition_training_notes": trans_full or None,
            "primary_discipline_count": _i(ws.cell(row=r, column=10).value),
            "secondary_discipline_count": _i(ws.cell(row=r, column=11).value),
            "status_label": _t(ws.cell(row=r, column=12).value),
            "constituent_movements": movements,
            "endurance_profile": endurance,
            "participation_format": participation,
            "multi_discipline": multi_disc,
        })
    return rows


def extract_disciplines(ws: Worksheet) -> list[dict[str, Any]]:
    """Sheet 2 — Discipline Library. Header R1, data R2+."""
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        did = _t(ws.cell(row=r, column=1).value)
        if not did or not did.startswith("D-"):
            continue
        min_base_text = _t(ws.cell(row=r, column=5).value)
        weeks_low, weeks_high = _parse_weeks(min_base_text)
        age_text = _t(ws.cell(row=r, column=8).value)
        age = _parse_age_ramps(age_text)

        recovery_text = _t(ws.cell(row=r, column=12).value)
        rows.append({
            "discipline_id": did,
            "discipline_name": _t(ws.cell(row=r, column=2).value) or did,
            # Source col 3 was the free-text terrain `discipline_category`,
            # removed (superseded by the curated `endurance_profile` the
            # discipline canon stamps on in normalize_dimension_rows).
            # No source column on the v11 sheet; populated post-hoc via
            # migrate_disciplines_add_primary_movement_v1.sql (like
            # stimulus_components). Value ∈ ENUM_MOVEMENTS when set.
            "primary_movement": None,
            "min_base_phase_text": min_base_text,
            "min_base_phase_weeks_low": weeks_low,
            "min_base_phase_weeks_high": weeks_high,
            "periodization_text": _t(ws.cell(row=r, column=6).value),
            "ramp_text": _t(ws.cell(row=r, column=7).value),
            "age_adjusted_ramp_text": age_text,
            "age_ramp_40_44_pct": age["40_44"],
            "age_ramp_45_54_pct": age["45_54"],
            "age_ramp_55_plus_pct": age["55_plus"],
            "taper_norms_text": _t(ws.cell(row=r, column=9).value),
            "common_injury_patterns": _split_dot_list(ws.cell(row=r, column=10).value),
            "injury_preceding_behaviors": _split_dot_list(ws.cell(row=r, column=11).value),
            "recovery_priority_text": recovery_text,
            "recovery_modalities": _parse_recovery_modalities(recovery_text),
            "evidence_quality_text": _t(ws.cell(row=r, column=13).value),
            # v10: schema column added for future use; v10 sheet has no source col yet
            "stimulus_components": None,
        })
    return rows


def extract_sport_discipline_map(
    ws: Worksheet,
    dropped_dupes: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Sheet 3 — Sport × Discipline Map. Header R1, data R2+.

    Dedupes by `(sport_name, discipline_id)` to satisfy the spec's UNIQUE
    constraint. First-seen wins. The source has three known duplicate
    keys as of v6 — one true duplicate (Triathlon D-002 listed twice
    identically) and two genuine sub-format splits where multiple
    disciplines share a `discipline_id` (Long Distance / Endurance
    Cycling D-006 and D-008). Dropped rows are appended to
    `dropped_dupes` if provided, for surfacing in the report.
    """
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for r in range(2, ws.max_row + 1):
        sport = _t(ws.cell(row=r, column=1).value)
        did = _t(ws.cell(row=r, column=2).value)
        if not sport or not did:
            continue
        race_text = _t(ws.cell(row=r, column=6).value)
        race_low, race_high = _parse_race_time_pct(race_text)
        row = {
            "sport_name": sport,
            "discipline_id": did,
            "discipline_name": _t(ws.cell(row=r, column=3).value) or did,
            "applicability": _t(ws.cell(row=r, column=4).value) or "INCLUDED",
            "role": _t(ws.cell(row=r, column=5).value) or "",
            "race_time_pct_text": race_text,
            "race_time_pct_low": race_low,
            "race_time_pct_high": race_high,
            "sport_specific_context": _t(ws.cell(row=r, column=7).value),
            "b2b_pairing_rule_text": _t(ws.cell(row=r, column=8).value),
            "phase_load_text": _t(ws.cell(row=r, column=9).value),
        }
        key = (sport, did)
        if key in seen:
            if dropped_dupes is not None:
                dropped_dupes.append({
                    "row_number": r,
                    "sport_name": sport,
                    "discipline_id": did,
                    "discipline_name": row["discipline_name"],
                    "role": row["role"],
                })
            continue
        seen.add(key)
        rows.append(row)
    return rows


def extract_discipline_pairing_matrix(
    ws: Worksheet,
    *,
    debug_meta: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """Sheet 4 — primary matrix. Header on R10; data R11+.

    The last data row is detected dynamically: scan downward until the
    first row whose col 1 doesn't start with `D-`. Rows after the matrix
    (KEY PAIRING RATIONALE etc.) are commentary, not pairing rows.

    Header cells starting with "D-" yield a destination column. Cells whose
    first token doesn't match `D-\\d+[a-z]?` are non-discipline columns and
    are ignored. (The `[a-z]?` suffix support is retained for robustness
    though the v11 R6 renumber removed all suffix ids.)
    """
    rows: list[dict[str, Any]] = []
    rating_map = {
        "PRE": "PREFERRED",
        "ACC": "ACCEPTABLE",
        "AVO": "AVOID",
        "IMP": "IMPRACTICAL",
        "N/A": "N/A",
        "PREFERRED": "PREFERRED",
        "ACCEPTABLE": "ACCEPTABLE",
        "AVOID": "AVOID",
        "IMPRACTICAL": "IMPRACTICAL",
    }

    _ID_RE = re.compile(r"(D-\d+[a-z]?)")

    # Read header row R10 to extract destination D-IDs
    header_ids: list[str | None] = [None]  # col 1 is "FROM" label
    for c in range(2, ws.max_column + 1):
        val = _t(ws.cell(row=10, column=c).value)
        m = _ID_RE.search(val or "")
        header_ids.append(m.group(1) if m else None)

    # Dynamically find the last data row: first empty/non-D row signals end.
    last_data_row = 10
    for r in range(11, ws.max_row + 1):
        first = _t(ws.cell(row=r, column=1).value)
        if not first or not _ID_RE.match(first):
            break
        last_data_row = r

    # Dedupe by (from_id, to_id), first-seen-wins, and skip self-pairs. The
    # R6 craft collapse maps two former ids onto one survivor (D-008a/b →
    # D-010, D-022/3 → D-024), so the survivor appears in two header columns
    # and two from-rows — without this the scanner emits the same canonical
    # pair twice (UniqueViolation on the discipline_pairing load) and turns the
    # old off-diagonal cell into a meaningless self-pair (D-010, D-010).
    seen: set[tuple[str, str]] = set()
    for r in range(11, last_data_row + 1):
        first = _t(ws.cell(row=r, column=1).value)
        if not first:
            continue
        m = _ID_RE.match(first)
        if not m:
            continue
        from_id = m.group(1)
        for c in range(2, ws.max_column + 1):
            to_id = header_ids[c - 1]
            if not to_id:
                continue
            if from_id == to_id:
                continue
            if (from_id, to_id) in seen:
                continue
            cell = _t(ws.cell(row=r, column=c).value)
            if not cell:
                continue
            rating = rating_map.get(cell.upper(), cell.upper())
            seen.add((from_id, to_id))
            rows.append({
                "discipline_id_a": from_id,
                "discipline_id_b": to_id,
                "pairing_rating": rating,
                "rationale": None,
                "source": "matrix",
            })

    if debug_meta is not None:
        debug_meta["matrix_last_data_row"] = last_data_row
        debug_meta["matrix_header_ids"] = [h for h in header_ids if h]
    return rows


def extract_pairing_b2b_fallback(
    sport_discipline_rows: list[dict[str, Any]],
    discipline_name_to_id: dict[str, str],
    matrix_pairs: set[tuple[str, str]],
) -> list[dict[str, Any]]:
    """Parse Sheet 3 col 7 `b2b_pairing_rule_text` per spec §4.6 (3).

    Each line is `→ {discipline_name}: {RATING}` or
    `→ {discipline_name}: {RATING} ({rationale})`.

    Skips pairs already present in the matrix (matrix wins).
    """
    rows: list[dict[str, Any]] = []
    pat = re.compile(
        r"→\s*(?P<dest>[^:]+?)\s*:\s*"
        r"(?P<rating>PREFERRED|ACCEPTABLE|AVOID|IMPRACTICAL|N/A)"
        r"(?:\s*\((?P<rat>[^)]+)\))?",
        re.IGNORECASE,
    )
    seen: set[tuple[str, str]] = set()

    # build name→id map case-insensitive
    name_lookup = {k.lower(): v for k, v in discipline_name_to_id.items()}

    for sdrow in sport_discipline_rows:
        from_id = sdrow["discipline_id"]
        text = sdrow.get("b2b_pairing_rule_text")
        if not text or not from_id:
            continue
        for line in str(text).splitlines():
            m = pat.search(line)
            if not m:
                continue
            dest_name = m.group("dest").strip()
            # Strip trailing parens or notes from the dest name
            dest_clean = re.sub(r"\s*\([^)]*\)\s*$", "", dest_name).strip()
            dest_id = name_lookup.get(dest_clean.lower())
            if not dest_id:
                continue
            key = (from_id, dest_id)
            if key in matrix_pairs or key in seen:
                continue
            seen.add(key)
            rating = m.group("rating").upper()
            rationale = m.group("rat")
            rows.append({
                "discipline_id_a": from_id,
                "discipline_id_b": dest_id,
                "pairing_rating": rating,
                "rationale": rationale.strip() if rationale else None,
                "source": "b2b_rule",
            })
    return rows


def extract_phase_load_allocation(
    ws: Worksheet,
    *,
    split_stats: dict[str, int] | None = None,
) -> list[dict[str, Any]]:
    """Sheet 5 — Phase Load Allocation. Header R1, data R2+.

    v10 additions:
      - Col 14 `Default Inclusion` (enum: included / excluded / prompt_required).
      - Notes (col 13) is split into:
          * `prescription_note` — leading non-audit clause (≤120 chars)
          * `audit_log` — remaining audit / source / patch chunks
          * `raw_notes` — the original cell text, unaltered
        See `_split_phase_load_notes` for the heuristic. Counts of rows
        producing a non-NULL `prescription_note` are returned via `split_stats`
        if provided.
    """
    rows: list[dict[str, Any]] = []
    last_sport: str | None = None
    n_with_prescription = 0
    n_total = 0
    for r in range(2, ws.max_row + 1):
        sport = _t(ws.cell(row=r, column=1).value)
        if sport:
            last_sport = sport
        disc = _t(ws.cell(row=r, column=3).value)
        role = _t(ws.cell(row=r, column=4).value)
        if not disc and not role:
            continue

        raw_notes = _t_raw(ws.cell(row=r, column=13).value)
        prescription, audit = _split_phase_load_notes(raw_notes)
        default_inclusion = _t(ws.cell(row=r, column=14).value)
        # Normalize whitespace; preserve original token for validation
        if default_inclusion:
            default_inclusion = default_inclusion.strip()

        n_total += 1
        if prescription:
            n_with_prescription += 1

        rows.append({
            "sport_name": last_sport or "",
            "discipline_id": _t(ws.cell(row=r, column=2).value),
            "discipline_name": disc or "",
            "role": role or "",
            "base_pct_low": _f(ws.cell(row=r, column=5).value),
            "base_pct_high": _f(ws.cell(row=r, column=6).value),
            "build_pct_low": _f(ws.cell(row=r, column=7).value),
            "build_pct_high": _f(ws.cell(row=r, column=8).value),
            "peak_pct_low": _f(ws.cell(row=r, column=9).value),
            "peak_pct_high": _f(ws.cell(row=r, column=10).value),
            "taper_pct_low": _f(ws.cell(row=r, column=11).value),
            "taper_pct_high": _f(ws.cell(row=r, column=12).value),
            "notes_conditions": _t(ws.cell(row=r, column=13).value),
            "default_inclusion": default_inclusion,
            "prescription_note": prescription,
            "audit_log": audit,
            "raw_notes": raw_notes,
        })

    if split_stats is not None:
        split_stats["rows"] = n_total
        split_stats["with_prescription"] = n_with_prescription
    return rows


def extract_phase_load_weekly_totals(
    ws: Worksheet,
    *,
    parse_failures: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Spec v3 §4.6b — derive 4 (Base/Build/Peak/Taper) hour-range rows per
    sport from the `WEEKLY TOTAL TARGET` aggregator row in Phase Load
    Allocation. Parse the Notes (col 13) cell for `PHASE: low–high hrs`
    matches.

    Discipline-name detection is case-insensitive and matches both em-dash
    `WEEKLY TOTAL TARGET` and minor typo variants. Rows whose Notes cell
    fails to yield 4 phase ranges contribute zero output rows and are
    appended to `parse_failures` if provided.
    """
    rows: list[dict[str, Any]] = []
    last_sport: str | None = None
    for r in range(2, ws.max_row + 1):
        sport = _t(ws.cell(row=r, column=1).value)
        if sport:
            last_sport = sport
        disc = _t(ws.cell(row=r, column=3).value)
        if not disc or "WEEKLY TOTAL TARGET" not in disc.upper():
            continue
        notes = _t_raw(ws.cell(row=r, column=13).value)
        parsed = _parse_weekly_total_text(notes)
        if not parsed:
            if parse_failures is not None:
                parse_failures.append({
                    "row_number": r,
                    "sport_name": last_sport or "",
                    "weekly_target_text": notes,
                })
            continue
        for phase in ("Base", "Build", "Peak", "Taper"):
            low, high, unit = parsed[phase]
            # The `weekly_low_hours` / `weekly_high_hours` column names are
            # legacy (the schema predates the km-volume open-water swim
            # sports). They hold the numeric range in whatever unit the
            # source row used; `weekly_unit` ('hrs' | 'km') disambiguates.
            # Consumers MUST check `weekly_unit` before treating the values
            # as hours.
            rows.append({
                "sport_name": last_sport or "",
                "phase": phase,
                "weekly_low_hours": low,
                "weekly_high_hours": high,
                "weekly_target_text": notes,
                "weekly_unit": unit,
            })
    return rows


_TEAM_FORMAT_SKIP_PREFIXES = ("PARADIGM", "FORMAT KEY", "TRAINING IMPLICATION")


def extract_team_formats(ws: Worksheet) -> list[dict[str, Any]]:
    """Sheet 6 — header on R3, data R4+. Skip paradigm separator rows."""
    rows: list[dict[str, Any]] = []
    for r in range(4, ws.max_row + 1):
        sport = _t(ws.cell(row=r, column=1).value)
        if not sport:
            continue
        # Skip section headers / paradigm separators
        upper = sport.upper()
        if any(upper.startswith(p) for p in _TEAM_FORMAT_SKIP_PREFIXES):
            continue
        rows.append({
            "sport_name": sport,
            "formats_available": _t(ws.cell(row=r, column=2).value),
            "team_format_types": _t(ws.cell(row=r, column=3).value),
            "unified_team_description": _t(ws.cell(row=r, column=4).value),
            "relay_specialist_description": _t(ws.cell(row=r, column=5).value),
            "training_implication_unified": _t(ws.cell(row=r, column=6).value),
            "training_implication_relay": _t(ws.cell(row=r, column=7).value),
            "key_distinctions_notes": _t(ws.cell(row=r, column=8).value),
        })
    return rows


def extract_cross_sport_properties(ws: Worksheet) -> list[dict[str, Any]]:
    """Sheet 8 — header R1, data R2+.

    v10 schema (9 cols): Property ID | Property Name | Description | Scope |
    Ranking | Estimated Values | Source(s) | Confidence | Notes.

    Filters by Property ID regex (`^[A-Z]+_[A-Z]+_\\d{3}$`) so EXTENSION
    NOTES commentary, blank rows, and banner cells are skipped automatically.
    """
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        prop = _t(ws.cell(row=r, column=1).value)
        if not prop or not _PROPERTY_ID_RE.match(prop.strip()):
            continue
        rows.append({
            "property_id": prop.strip(),
            "property_name": _t(ws.cell(row=r, column=2).value) or prop.strip(),
            "description": _t(ws.cell(row=r, column=3).value),
            "scope": _t(ws.cell(row=r, column=4).value),
            "ranking_text": _t(ws.cell(row=r, column=5).value),
            "estimated_values": _t(ws.cell(row=r, column=6).value),
            "source_text": _t(ws.cell(row=r, column=7).value),
            # `source_evidence` retained for backwards compatibility with the
            # v2 schema column; same content as source_text.
            "source_evidence": _t(ws.cell(row=r, column=7).value),
            "confidence": _t(ws.cell(row=r, column=8).value),
            "notes": _t(ws.cell(row=r, column=9).value),
        })
    return rows


def extract_discipline_substitutes(
    wb,
    *,
    parse_warnings: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Spec v3 §4.13 — read the Discipline Substitution Map sheet.

    Columns (R1 header, R2+ data): Target ID | Target Name | Substitute ID |
    Substitute Name | Fidelity (0-1) | Constraints | Category.

    Rows with a missing or non-numeric `fidelity` are dropped and (if
    `parse_warnings` is provided) appended for the report. `substitute_covers`
    is initialized to NULL — populated in a later session.
    """
    if "Discipline Substitution Map" not in wb.sheetnames:
        return []
    ws = wb["Discipline Substitution Map"]
    rows: list[dict[str, Any]] = []
    # Dedupe by (target_id, substitute_id, substitute_name), first-seen-wins,
    # and skip self-substitutes. The R6 craft collapse maps two former ids onto
    # one survivor (D-008a/b → D-010), so two source rows can collapse to the
    # same key (e.g. (D-010, D-011, 'Canoeing')) — without this the load trips
    # the UNIQUE(target_id, substitute_id, substitute_name, etl_version).
    seen: set[tuple[str, str, str]] = set()
    for r in range(2, ws.max_row + 1):
        target_id = _t(ws.cell(row=r, column=1).value)
        if not target_id:
            continue
        substitute_id = _t(ws.cell(row=r, column=3).value)
        if not substitute_id:
            continue
        if target_id == substitute_id:
            continue
        fidelity = _f(ws.cell(row=r, column=5).value)
        if fidelity is None:
            if parse_warnings is not None:
                parse_warnings.append({
                    "row_number": r,
                    "target_id": target_id,
                    "substitute_id": substitute_id,
                    "reason": "missing or non-numeric fidelity",
                })
            continue
        substitute_name = _t(ws.cell(row=r, column=4).value) or substitute_id
        key = (target_id, substitute_id, substitute_name)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            "target_id": target_id,
            "target_name": _t(ws.cell(row=r, column=2).value) or target_id,
            "substitute_id": substitute_id,
            "substitute_name": substitute_name,
            "fidelity": fidelity,
            "constraints": _t_raw(ws.cell(row=r, column=6).value),
            "category": _t(ws.cell(row=r, column=7).value),
            "substitute_covers": None,
        })
    return rows


def extract_discipline_training_gaps(wb) -> list[dict[str, Any]]:
    """Spec v3 §4.14 — read the Discipline Training Gaps sheet.

    Columns (R1 header, R2+ data): Discipline ID | Discipline Name |
    Gap Description (free text). Maps the free-text into structured
    `gap_type` and `multi_substitute_candidate` fields.
    """
    if "Discipline Training Gaps" not in wb.sheetnames:
        return []
    ws = wb["Discipline Training Gaps"]
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        did = _t(ws.cell(row=r, column=1).value)
        if not did:
            continue
        notes = _t_raw(ws.cell(row=r, column=3).value) or ""
        notes_l = notes.lower()
        if "no good single" in notes_l or "no good single substitute" in notes_l:
            gap_type = "no_single_substitute"
        elif (
            "no off-snow" in notes_l
            or "no off-environment" in notes_l
            or "no off-" in notes_l
        ):
            gap_type = "no_off_environment_substitute"
        elif "no discipline-level substitute" in notes_l:
            gap_type = "no_substitute_available"
        else:
            gap_type = "other"
        multi_sub = (
            "multi-substitute" in notes_l
            or "multi substitute" in notes_l
            or "compose" in notes_l
        )
        rows.append({
            "discipline_id": did,
            "discipline_name": _t(ws.cell(row=r, column=2).value) or did,
            "gap_type": gap_type,
            "notes": notes or None,
            "multi_substitute_candidate": multi_sub,
        })
    return rows


def build_sport_discipline_bridge(
    sport_discipline_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Spec §4.9 — derived bridge. One row per (sport, discipline_id) where
    applicability='INCLUDED'.

    `exercise_db_sport`: defaults to the framework `sport_name`. The vocab
    alignment validator (Open Item #5) surfaces mismatches between this
    field and `sport_exercise_map.sport_name` for manual reconciliation.
    """
    rows: list[dict[str, Any]] = []
    for r in sport_discipline_rows:
        if (r.get("applicability") or "").upper() != "INCLUDED":
            continue
        rows.append({
            "framework_sport": r["sport_name"],
            "discipline_id": r["discipline_id"],
            "discipline_name": r["discipline_name"],
            "exercise_db_sport": r["sport_name"],
            "role": r["role"],
            "default_race_time_pct_low": r.get("race_time_pct_low"),
            "default_race_time_pct_high": r.get("race_time_pct_high"),
        })
    return rows
