"""export_xlsx — DB→xlsx bulk-review export for Layer 0.

The DB-as-source-of-truth model (`Layer0_AuthoringModel_DBSourceOfTruth_Design
_v1.md`) retires the authoring workbooks. The one thing the spreadsheet was
genuinely better at — seeing every exercise / sport / vocabulary row at once for
bulk review — is recovered by this read-only export (design §3.5 / §6.5, decision
B, Andy 2026-06-10). It is a *projection* of the live DB, never an authoring
input: edits still arrive as `etl/migrations/layer0/` SQL.

One sheet per `layer0.*` table, active rows only (`WHERE superseded_at IS NULL`).
Tables are discovered from `information_schema`, so a new table is picked up with
no list to maintain here.

Mirrors `validate_layer0.py`: the serialization + workbook assembly is pure and
unit-tested DB-free (`etl/tests/test_export_xlsx.py`); `collect()` is the only
DB-touching step; `main()` imports `db` lazily so the pure logic carries no
psycopg2 / live-connection dependency.

Usage:
    python -m etl.layer0.export_xlsx                 # reads DATABASE_URL
    python -m etl.layer0.export_xlsx --out path.xlsx

Exit code: 0 on success.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook

DEFAULT_OUT = Path(__file__).parent.parent / "output" / "layer0_db_export.xlsx"

# openpyxl rejects these in sheet titles.
_BAD_SHEET_CHARS = set(r"[]:*?/\\")
_MAX_SHEET_NAME = 31


# ─── Pure serialization (unit-tested DB-free) ─────────────────────────


def cell_value(value: Any) -> Any:
    """Render one DB value as an xlsx-cell-friendly scalar.

    psycopg2 hands back native Python types: TEXT[] → list, JSONB → dict/list,
    NUMERIC → Decimal, TIMESTAMPTZ → tz-aware datetime. openpyxl writes str /
    int / float / bool natively but *raises* on tz-aware datetimes and does not
    accept Decimal, lists, or dicts — so those are flattened to readable text.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float, str)):
        return value
    if isinstance(value, list):
        # TEXT[] reads cleanest comma-joined; a JSONB array of objects falls
        # through to JSON so nothing is lost.
        if all(isinstance(v, str) for v in value):
            return ", ".join(value)
        return json.dumps(value, default=str)
    if isinstance(value, dict):
        return json.dumps(value, default=str)
    return str(value)


def safe_sheet_name(name: str, used: set[str]) -> str:
    """An openpyxl-legal, unique sheet title (≤31 chars, no reserved chars)."""
    cleaned = "".join("_" if c in _BAD_SHEET_CHARS else c for c in name)
    cleaned = cleaned[:_MAX_SHEET_NAME] or "sheet"
    candidate = cleaned
    n = 1
    while candidate in used:
        suffix = f"_{n}"
        candidate = cleaned[: _MAX_SHEET_NAME - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def build_workbook(tables: dict[str, tuple[list[str], list]]) -> Workbook:
    """Assemble the review workbook: one sheet per table, header frozen.

    `tables` maps table name → (column names, rows). Iteration order is
    preserved (callers pass alphabetically-discovered tables).
    """
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    used: set[str] = set()
    for table, (columns, rows) in tables.items():
        ws = wb.create_sheet(safe_sheet_name(table, used))
        ws.append(list(columns))
        ws.freeze_panes = "A2"
        for row in rows:
            ws.append([cell_value(v) for v in row])
    if not wb.sheetnames:  # a layer0 schema with zero tables — keep it saveable
        wb.create_sheet("layer0")
    return wb


# ─── DB read (the only DB-touching step) ──────────────────────────────


def discover_tables(conn) -> list[str]:
    """All base tables in the `layer0` schema, alphabetical."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT table_name
              FROM information_schema.tables
             WHERE table_schema = 'layer0'
               AND table_type = 'BASE TABLE'
             ORDER BY table_name
            """
        )
        return [r[0] for r in cur.fetchall()]


def fetch_table(conn, table: str) -> tuple[list[str], list]:
    """Active rows of one layer0 table → (column names, rows). Filters
    `superseded_at IS NULL` when the table carries that column."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT 1 FROM information_schema.columns
             WHERE table_schema = 'layer0' AND table_name = %s
               AND column_name = 'superseded_at'
            """,
            (table,),
        )
        where = " WHERE superseded_at IS NULL" if cur.fetchone() else ""
        cur.execute(f'SELECT * FROM layer0."{table}"{where} ORDER BY 1')
        columns = [d[0] for d in cur.description]
        return columns, cur.fetchall()


def collect(conn) -> dict[str, tuple[list[str], list]]:
    """Read every layer0 table's active rows. The only DB-touching step."""
    return {table: fetch_table(conn, table) for table in discover_tables(conn)}


# ─── Main ─────────────────────────────────────────────────────────────


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Export the live layer0.* tables to an xlsx for bulk review."
    )
    parser.add_argument(
        "--out",
        default=str(DEFAULT_OUT),
        help=f"Output .xlsx path (default: {DEFAULT_OUT}).",
    )
    args = parser.parse_args(argv)

    # Lazy so the pure logic above (and its unit tests) need no psycopg2.
    from etl.layer0 import db

    with db.connect() as conn:
        tables = collect(conn)
    wb = build_workbook(tables)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    total_rows = sum(len(rows) for _, rows in tables.values())
    print(
        f"Wrote {out_path} — {len(tables)} table(s), {total_rows} active row(s), "
        f"{out_path.stat().st_size:,} bytes"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
