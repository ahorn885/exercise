#!/usr/bin/env python3
"""Generate the v4 vocabulary-reconciliation workbook (Google-Sheets-importable .xlsx)."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEAD = Font(bold=True, color="FFFFFF")
HEADFILL = PatternFill("solid", fgColor="305496")
WRAP = Alignment(vertical="top", wrap_text=True)

def add_sheet(name, headers, rows, widths):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD
        cell.fill = HEADFILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    ws.freeze_panes = "A2"
    return ws

# ── README ────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "README"
readme = [
 ["Vocabulary Reconciliation — v4", ""],
 ["Date", "2026-06-08"],
 ["", ""],
 ["Cell conventions", ""],
 ["full value", "the actual value, in full (no slugs without their label)"],
 ["N/A", "this column does not apply to this row's domain"],
 ["missing", "this column APPLIES to this row but the value is absent there — a real gap to reconcile"],
 ["", ""],
 ["Source legend", ""],
 ["layer0 live", "the live layer0.* ETL tables (v1.5.0) — proposed source of truth"],
 ["v1 EQUIPMENT_CATEGORIES", "init_db.py:2132 — the v1 app picker (slug + label)"],
 ["v1 athlete enum", "athlete.py — PADDLE_CRAFT_TYPES / bike_types_available"],
 ["Committed elsewhere", "a firm decision or proposal from a GitHub issue or an aidstation-sources spec"],
 ["", ""],
 ["Sources mined", "live layer0.* tables + all 100 open/closed GitHub issues + all aidstation-sources specs & handoffs"],
 ["Tabs", "Disciplines | Craft_Vessels | Equipment_General | Terrain | Modality_Groups | Decisions"],
 ["Note", "If a decision exists only in conversation (not written in an issue/doc), it is not here — point me at it."],
]
for r in readme: ws.append(r)
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 90
ws["A1"].font = Font(bold=True, size=14)
for rr in (4, 9):
    ws.cell(row=rr, column=1).font = Font(bold=True)
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = WRAP

# ── Disciplines ───────────────────────────────────────────────────────────
add_sheet("Disciplines",
 ["ID", "Name", "Modality group(s)", "In layer0 live?", "Committed change", "Source / ref", "Status", "Notes"],
 [
  ["D-001","Trail Running","foot","yes","none","—","live",""],
  ["D-002","Road Running","foot","yes","none","—","live",""],
  ["D-003","Trekking","foot","yes","absorbed D-015 Orienteering","discipline_canon","live","renamed from Hiking"],
  ["D-004","Swimming","swim_openwater","yes","absorbed D-005 + D-016","discipline_canon","live",""],
  ["D-006","Road Cycling","bike_pavement","yes","split → D-006a / D-006b / D-006c","#477 (FIRM)","committed-pending-ETL","Endurance-Cycling bridge; global vs bridge-only TBD (D3)"],
  ["D-007","Time-Trial Cycling","bike_pavement","yes","REMOVE → folds into D-006c","#476 / #477 (FIRM)","committed-pending-ETL",""],
  ["D-008","Mountain Biking","bike_offroad","yes","split → D-008a / D-008b","#477 (FIRM)","committed-pending-ETL",""],
  ["D-009","Packrafting","paddle_flatwater; paddle_whitewater","yes","none","—","live","whitewater incl. (Andy confirmed)"],
  ["D-010","Kayaking","paddle_flatwater; paddle_whitewater","yes","D-010a / D-010b flatwater/whitewater split","UNVERIFIED — no issue/doc","open-question","D4 — confirm decided or drop"],
  ["D-011","Canoeing","paddle_flatwater","yes","none","—","live",""],
  ["D-012","Rock Climbing","climb","yes","none","—","live",""],
  ["D-013","Abseiling","climb","yes","none","—","live",""],
  ["D-014","Via Ferrata","climb","yes","none","—","live",""],
  ["D-017","Snowshoeing","snow_travel","yes","none","—","live",""],
  ["D-018","Mountaineering","snow_travel","yes","TRN-005 vs TRN-012 routing","Synthesis_Design_v2 §5.1 (OPEN)","open-question","D7 — routing semantics"],
  ["D-019","Paddle Rafting","paddle_flatwater","yes","none","—","live",""],
  ["D-021","Uphill Skinning","snow_glide","yes","none","—","live",""],
  ["D-022","Alpine Descent","snow_glide","yes","none","—","live",""],
  ["D-024","Mountain Running","foot","yes","none","—","live",""],
  ["D-027","Obstacle Course Racing","foot","yes","none","—","live",""],
  ["D-028","Cross-Country Skiing","snow_glide","yes","none","—","live",""],
  ["D-006a","Road Cycling (Long Distance)","bike_pavement","missing","NEW","#477 (FIRM)","committed-pending-ETL","replaces bridge r58"],
  ["D-006b","Road Cycling (Gravel)","bike_offroad","missing","NEW","#477 (FIRM)","committed-pending-ETL","replaces r59"],
  ["D-006c","Time Trial Cycling","bike_pavement","missing","NEW; absorbs D-007","#477 (FIRM)","committed-pending-ETL","replaces r60 + D-007"],
  ["D-008a","Mountain Biking (XC)","bike_offroad","missing","NEW","#477 (FIRM)","committed-pending-ETL","replaces r61"],
  ["D-008b","Mountain Biking (Enduro)","bike_offroad","missing","NEW","#477 (FIRM)","committed-pending-ETL","replaces r62"],
  ["D-015","Orienteering","foot (via D-003)","missing (merged → D-003)","merged","discipline_canon (#320)","removed-via-canon",""],
  ["D-005","Pool Sprint Swimming","swim (via D-004)","missing (merged → D-004)","merged","discipline_canon","removed-via-canon",""],
  ["D-016","Generic Swimming","swim (via D-004)","missing (merged → D-004)","merged","discipline_canon","removed-via-canon",""],
 ],
 [9, 28, 30, 16, 26, 28, 22, 44])

# ── Craft vessels ─────────────────────────────────────────────────────────
add_sheet("Craft_Vessels",
 ["Vessel", "Type", "layer0.equipment_items", "v1 EQUIPMENT_CATEGORIES", "v1 athlete enum", "Maps to discipline(s)", "Status", "Notes"],
 [
  ["Road bike","cycling","Road bike",'road_bike "Road Bike"',"road_bike (bike_types_available)","D-006 → D-006a","live",""],
  ["Mountain bike","cycling","Mountain bike",'mountain_bike "Mountain Bike (MTB)"',"mountain_bike (bike_types_available)","D-008 → D-008a + D-008b","live","Andy: mtb → mtb + XC"],
  ["Gravel bike","cycling","Gravel bike",'gravel_bike "Gravel Bike"',"gravel_bike (bike_types_available)","Road + XC (Andy: D-006a + D-008a)","live","Andy: gravel → road + XC, not MTB"],
  ["TT / triathlon bike","cycling","TT / triathlon bike","missing","missing","D-006c (after #477)","live (catalog only)","v1 picker can't select it"],
  ["Bike (generic)","cycling","Bike (generic)","missing","missing","D-006 (ambiguous)","prune-candidate","D2"],
  ["Cycling trainer (indoor)","cycling","Bike trainer",'cycling_trainer "Cycling Trainer / Smart Trainer"',"cycling_trainer (bike_types_available)","N/A (indoor, not a craft)","live","name mismatch: Bike trainer vs Cycling Trainer"],
  ["Kayak","paddle","Kayak",'kayak "Kayak"',"kayak (PADDLE_CRAFT_TYPES)","D-010","live",""],
  ["Canoe","paddle","Canoe",'canoe "Canoe"',"canoe (PADDLE_CRAFT_TYPES)","D-011","live",""],
  ["Packraft","paddle","Packraft",'packraft "Packraft"',"packraft (PADDLE_CRAFT_TYPES)","D-009","live","whitewater incl."],
  ["Surfski","paddle","missing","missing","surfski (PADDLE_CRAFT_TYPES)","no discipline","prune-candidate","Andy: never tracked"],
  ["Sea kayak","paddle","Sea kayak","missing","missing","D-010","prune-candidate","Andy: never tracked; ETL readers: sum_to_100, report.py"],
  ["SUP","paddle","SUP","missing","missing","no discipline","prune-candidate","orphan (no SUP discipline)"],
  ["Inflatable raft","paddle","Inflatable raft","missing","missing","D-019","prune-candidate","Andy: never tracked"],
  ["Rowing shell","paddle","Rowing shell","missing","missing","no discipline","prune-candidate","orphan (no rowing discipline)"],
 ],
 [22, 9, 24, 38, 34, 32, 18, 44])

# ── General equipment (A vs C) ────────────────────────────────────────────
gen = [
 ["Barbell","Barbells & Bars",'barbell "Barbell (Olympic)"',"both",""],
 ["EZ curl bar","Barbells & Bars",'ez_bar "EZ Curl Bar"',"both",""],
 ["Trap bar","Barbells & Bars",'hex_bar "Hex / Trap Bar"',"both",""],
 ["Dumbbell","Dumbbells",'dumbbells "Dumbbells"',"both",""],
 ["Kettlebell","Kettlebells",'kettlebell "Kettlebell"',"both",""],
 ["Sandbag","Plyo & Power",'sandbag "Sandbag"',"both",""],
 ["Medicine ball","Plyo & Power",'med_ball "Med Ball"',"both",""],
 ["Squat rack","Barbells & Bars",'squat_rack "Squat Rack / Power Cage"',"both",""],
 ["Smith machine","Barbells & Bars",'smith_machine "Smith Machine"',"both",""],
 ["Bench","Bodyweight & Portable Equipment",'bench_flat "Flat Bench" + bench_adjustable "Adjustable / Incline Bench"',"both","C merges two A rows"],
 ["Glute ham developer (GHD)","Machines — Lower Body",'ghd "GHD / Hyperextension Bench"',"both",""],
 ["Pull-up bar","Bodyweight & Portable Equipment",'pull_up_bar "Pull-Up Bar"',"both",""],
 ["Dip bars","Bodyweight & Portable Equipment",'dip_bars "Dip Bars / Parallel Bars"',"both",""],
 ["Gymnastic rings","Bodyweight & Portable Equipment",'rings "Gymnastic Rings"',"both",""],
 ["Leg press machine","Machines — Lower Body",'leg_press "Leg Press"',"both",""],
 ["Hack squat machine","Machines — Lower Body",'hack_squat "Hack Squat Machine"',"both",""],
 ["Leg extension machine","Machines — Lower Body",'leg_extension "Leg Extension Machine"',"both",""],
 ["Leg curl machine","Machines — Lower Body",'leg_curl "Leg Curl Machine"',"both",""],
 ["Standing calf raise machine","Machines — Lower Body",'calf_raise_machine "Calf Raise Machine"',"both",""],
 ["Cable machine","Machines — Upper Body",'cable_machine "Cable Machine / Crossover"',"both",""],
 ["Lat pulldown machine","Machines — Upper Body",'lat_pulldown "Lat Pulldown Machine"',"both",""],
 ["Seated row machine","Machines — Upper Body",'seated_row_machine "Seated Row Machine"',"both",""],
 ["Chest press machine","Machines — Upper Body",'pec_deck "Pec Deck / Chest Fly Machine"',"both","approx — pec deck ≈ chest fly"],
 ["Shoulder press machine","Machines — Upper Body",'shoulder_press_machine "Shoulder Press Machine"',"both",""],
 ["Assisted pull-up / dip machine","Machines — Upper Body",'assisted_pullup "Assisted Pull-Up / Dip Machine"',"both",""],
 ["Treadmill","Machines — Cardio",'treadmill "Treadmill"',"both",""],
 ["Elliptical","Machines — Cardio",'elliptical "Elliptical / Cross Trainer"',"both",""],
 ["Stationary bike","Machines — Cardio",'stationary_bike "Stationary Bike (Upright)" + recumbent_bike "Recumbent Bike"',"both","C merges two A rows"],
 ["Spin bike","Machines — Cardio",'spin_bike "Spin Bike / Peloton"',"both",""],
 ["Stair climber","Machines — Cardio",'stair_climber "Stair Climber / StepMill"',"both",""],
 ["Rowing ergometer","Machines — Upper Body",'rowing_erg "Rowing Erg (Concept2)"',"both",""],
 ["Paddle ergometer","Machines — Cardio",'kayak_erg "Kayak Ergometer"',"both","name diff"],
 ["Assault bike","Machines — Cardio",'air_bike "Air Bike / Assault Bike"',"both",""],
 ["Ski erg","Machines — Upper Body",'ski_erg "SkiErg"',"both",""],
 ["Weighted sled","Plyo & Power",'sled "Sled"',"both",""],
 ["Battle ropes","Plyo & Power",'battle_ropes "Battle Ropes"',"both",""],
 ["Plyo box","Plyo & Power",'plyo_box "Plyo Box"',"both",""],
 ["Resistance band","Bodyweight & Portable Equipment",'resistance_bands "Resistance Bands"',"both",""],
 ["TRX / suspension trainer","Bodyweight & Portable Equipment",'trx "TRX / Suspension Trainer"',"both",""],
 ["Weighted vest","Sport-Specific — Running & Hiking",'weighted_vest "Weighted Vest"',"both",""],
 ["Jump rope","Bodyweight & Portable Equipment",'jump_rope "Jump Rope"',"both",""],
 ["Stability ball","Stability & Balance",'stability_ball "Stability Ball"',"both",""],
 ["BOSU ball","Stability & Balance",'bosu "BOSU Ball"',"both",""],
 ["Ab wheel","Bodyweight & Portable Equipment",'ab_wheel "Ab Wheel"',"both",""],
 ["Foam roller","Bodyweight & Portable Equipment",'foam_roller "Foam Roller"',"both",""],
 ["Grip trainer","Grip & Forearm Specific",'grip_trainer "Grip Trainer (squeeze)"',"both",""],
 ["Rice bucket","Grip & Forearm Specific",'rice_bucket "Rice Bucket"',"both",""],
 ["Lacrosse ball","Bodyweight & Portable Equipment",'lacrosse_ball "Lacrosse Ball / Massage Ball"',"both",""],
 ["Hangboard","Grip & Forearm Specific",'hangboard "Hangboard"',"both",""],
 ["missing","N/A",'treadwall "Treadwall"',"A-only","fold into C before deleting A"],
 ["missing","N/A",'climbing_wall "Climbing Wall / Bouldering"',"A-only","C uses TRN-014 Climbing Gym terrain instead"],
 ["missing","N/A",'tricep_bar "Tricep Bar (W-bar)"',"A-only","fold into C before deleting A"],
 ["missing","N/A",'preacher_bench "Preacher Curl Bench"',"A-only","fold into C before deleting A"],
 ["missing","N/A",'slam_ball "Slam Ball"',"A-only","fold into C before deleting A"],
]
add_sheet("Equipment_General",
 ["Canonical name (layer0)", "Category (layer0)", "v1 EQUIPMENT_CATEGORIES", "Coverage", "Notes"],
 gen, [30, 32, 56, 12, 44])

# ── Terrain ───────────────────────────────────────────────────────────────
add_sheet("Terrain",
 ["ID", "Name", "In layer0 live?", "Committed change", "Source / ref", "Status", "Notes"],
 [
  ["TRN-001","Road / Paved","yes","none","—","live",""],
  ["TRN-002","Groomed Trail","yes","none","—","live",""],
  ["TRN-003","Technical Trail","yes","none","—","live",""],
  ["TRN-004","Hill / Rolling","yes","none","—","live",""],
  ["TRN-005","Mountain / Alpine","yes","none","—","live",""],
  ["TRN-006","Fell / Moorland","yes","none","—","live",""],
  ["TRN-007","Technical Rock","yes",'RENAME → "Technical Rock/Scree"',"Synthesis_Design_v2 §5.1 (FIRM); UI relabel #444","committed-pending-ETL","D6"],
  ["TRN-008","Pool","yes","none","—","live",""],
  ["TRN-009","Flat Water","yes","none","—","live",""],
  ["TRN-010","Ocean / Tidal","yes","none","—","live",""],
  ["TRN-011","Whitewater","yes","none","—","live",""],
  ["TRN-012","Snow / Winter Alpine","yes","none","—","live",""],
  ["TRN-013","Rock Wall (Outdoor)","yes","none","—","live",""],
  ["TRN-014","Climbing Gym","yes","race_eligible = FALSE (proposal)","#445 (OPEN)","open-question",""],
  ["TRN-015","Pump Track / Skills Course","yes","race_eligible = FALSE (proposal)","#445 (OPEN)","open-question",""],
  ["TRN-016","Indoor / Gym","yes","race_eligible = FALSE (proposal)","#445 (OPEN)","open-question",""],
  ["TRN-017","Moving Water","yes","none","—","live","NOTE: id already used — see off-trail conflict"],
  ["TRN-020","Gravel","yes","none","—","live",""],
  ["TBD","Off-Trail / Bush (trackless)","missing","NEW; race_eligible = TRUE","Synthesis_Design_v2 §5.1 (FIRM); #340","committed-pending-ETL","D6 — needs free id; CARRY_FORWARD wrongly calls it TRN-017 (taken by Moving Water)"],
 ],
 [10, 30, 16, 34, 40, 22, 50])

# ── Modality groups ───────────────────────────────────────────────────────
add_sheet("Modality_Groups",
 ["group_id", "Description", "group_kind", "Member disciplines (live)", "Status"],
 [
  ["paddle_flatwater","Flatwater paddle","paddle","D-009, D-010, D-011, D-019","live"],
  ["paddle_whitewater","Whitewater paddle","paddle","D-009, D-010","live"],
  ["foot","Foot (run / hike / nav)","foot","D-001, D-002, D-003, D-024, D-027","live"],
  ["bike_pavement","Bike on pavement","bike","D-006, D-007 (→ D-006a, D-006c)","live"],
  ["bike_offroad","Bike off-road","bike","D-008 (→ D-006b, D-008a, D-008b)","live"],
  ["snow_travel","Snow travel (foot)","snow","D-017, D-018","live"],
  ["snow_glide","Snow travel (gliding)","snow","D-021, D-022, D-028","live"],
  ["climb","Climbing (rope-protected)","climb","D-012, D-013, D-014","live"],
  ["swim_openwater","Open-water swim","swim","D-004","live"],
 ],
 [18, 28, 12, 36, 12])

# ── Decisions ─────────────────────────────────────────────────────────────
add_sheet("Decisions",
 ["#", "Decision", "Recommendation", "Your call"],
 [
  ["D1","Single source of truth for craft/equipment vocabulary","layer0.equipment_items (C)",""],
  ["D2","Prune list: surfski, Sea kayak, SUP, Inflatable raft, Rowing shell, Bike (generic) — keep / prune / fold each","prune surfski; fold Sea kayak→Kayak, Inflatable raft→Packraft; decide SUP/Rowing/generic",""],
  ["D3","#477 cycling split (D-006a/b/c, D-008a/b, remove D-007) ETL — when, and global vs bridge-only IDs","ETL before X1b.3b; decide scope",""],
  ["D4","D-010a / D-010b kayak flatwater/whitewater split — decided or dropped?","confirm or drop (unverified today)",""],
  ["D5","Retire EQUIPMENT_CATEGORIES (A) entirely — fold 5 A-only items into C, re-source picker from C","yes, as its own slice",""],
  ["D6","Terrain micro-slice: off-trail new id (TRN-017 is taken by Moving Water) + TRN-007 rename","bundle as one terrain ETL slice",""],
  ["D7","D-018 Mountaineering routing: TRN-005 vs TRN-012 (OR-match vs snow-only)","open — your call",""],
 ],
 [6, 60, 50, 30])

import os
out = "/home/user/exercise/aidstation-sources/vocab_reconciliation"
os.makedirs(out, exist_ok=True)
path = os.path.join(out, "Vocabulary_Reconciliation_v4.xlsx")
wb.save(path)
print("wrote", path)
print("sheets:", wb.sheetnames)
