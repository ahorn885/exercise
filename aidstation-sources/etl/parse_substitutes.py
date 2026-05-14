"""
parse_substitutes.py — v3 with AND-OR schema, Stairs vocab, Backpack as canonical.

Schema change: equipment_required is now list[list[str]] (CNF).
  - Outer list = OR (any group satisfies)
  - Inner list = AND (all items in group must be available)
  - Empty outer list ([]) = no equipment required

Examples:
  [["Dumbbell"]]                                  → needs Dumbbell
  [["Dumbbell"], ["Kettlebell"]]                  → needs Dumbbell OR Kettlebell
  [["Stairs", "Backpack"], ["Stairs", "Weighted vest"]]
                                                  → needs (Stairs AND Backpack) OR (Stairs AND Weighted vest)
  []                                              → no equipment required (bodyweight or fully improvised)
"""

import openpyxl, re, json
from pathlib import Path

# ── Hardcoded "real Stairs" entries (full flight required) ───────────────
# Identified by manual review; these get Stairs equipment vs other "stair"
# mentions which are universal step / edge / improvised.
REAL_STAIRS_ENTRIES = {
    # (ex_id, substring of substitute text) → True if needs full Stairs equipment
    ('EX010', 'Stair climb with pack'),
    ('EX048', 'Hotel or building stairs — walk/run up'),
    ('EX050', 'Stair climb with pack'),
    ('EX050', 'Hotel stairs with loaded backpack'),
    ('EX050', 'building stairwell with pack'),
    ('EX073', 'Hotel stairs at sustained effort'),
}

def is_real_stairs(ex_id: str, text: str) -> bool:
    return any(ex == ex_id and snippet in text for ex, snippet in REAL_STAIRS_ENTRIES)


# ── Equipment patterns ─────────────────────────────────────────────────────
EQUIPMENT_PATTERNS = [
    (r'^🏠', [], True),

    # Specific machines
    (r'\bSafety Bar\b', ['Safety squat bar'], False),
    (r'\bHack Squat\b', ['Hack squat machine'], False),
    (r'\bSmith Machine\b', ['Smith machine'], False),
    (r'\bLeg Press\b', ['Leg press machine'], False),
    (r'\bLeg Curl\b(?!.*Band)', ['Leg curl machine'], False),
    (r'\bLeg Extension\b', ['Leg extension machine'], False),
    (r'\bLat Pulldown\b', ['Lat pulldown machine'], False),
    (r'\bSeated Row\b|\bRow Machine\b|\bCable [Rr]ow\b', ['Seated row machine'], False),
    (r'\bCable\b', ['Cable machine'], False),
    (r'\bAssault Bike\b|\bAir Bike\b', ['Assault bike'], False),
    (r'\b(Indoor )?Row(ing)? Erg(ometer)?\b|\bRow Erg\b|\bConcept2\b', ['Rowing ergometer'], False),
    (r'\bArm Bike\b|\bUpper Body Erg(ometer)?\b|\bUBE\b', ['Arm bike / UBE'], False),
    (r'\bSki Erg\b|\bSki Ergometer\b|\bNordic [Ss]ki [Mm]achine\b|\bNordic [Ss]ki [Ee]rg(ometer)?\b', ['Ski erg'], False),
    (r'\bElliptical\b', ['Elliptical'], False),
    (r'\bStair Climber\b|\bStairmaster\b|\bStepmill\b', ['Stair climber'], False),
    (r'\bTreadmill\b', ['Treadmill'], False),
    (r'\bCycling Trainer\b|\bIndoor Trainer\b|\bBike Trainer\b', ['Bike trainer'], False),
    (r'\bGHD\b|\bGlute[- ]?Ham [Dd]eveloper\b', ['Glute ham developer (GHD)'], False),
    (r'\bHyperextension [Bb]ench\b|\bRoman [Cc]hair\b|\bReverse Hyper\b', ['Hyperextension bench'], False),
    (r'\bLandmine\b', ['Landmine attachment'], False),

    # Cross-training
    (r'\bRollerskis?\b', ['Rollerskis'], False),
    (r'\bInline [Ss]kates?\b', ['Inline skates'], False),

    # Explicit absence
    (r'\bno (dumbbell|DB|bench|band)\b|\bwithout (dumbbell|DB|bench|band)\b', [], False),
    (r'\bIT band\b|\biliotibial band\b', [], False),

    # Free weights
    (r'\bTrap Bar\b|\bHex Bar\b', ['Trap bar'], False),
    (r'\bEZ[- ]?[Bb]ar\b|\bEZ Curl Bar\b', ['EZ curl bar'], False),
    (r'\bDB\b|\bDumbbell', ['Dumbbell'], False),
    (r'\bKB\b|\bKettlebell', ['Kettlebell'], False),
    (r'\bBarbell\b|\bSumo deadlift\b', ['Barbell', 'Squat rack'], False),
    (r'\bSandbag\b', ['Sandbag'], False),
    (r'\b(Medicine [Bb]all|Med [Bb]all|Slam [Bb]all|Wall [Bb]all)\b', ['Medicine ball'], False),
    (r'\bPlate\b(?!let)', ['Weight plates'], False),
    (r'\bFat Gripz?\b', ['Fat grips'], False),

    # Backpack — NOW canonical equipment (not improvised env)
    (r'\b[Bb]ackpack\b', ['Backpack'], False),

    # Vest — already canonical
    (r'\bWeight(ed)? Vest\b', ['Weighted vest'], False),

    # Suspension / rings
    (r'\bTRX\b|\bSuspension Trainer\b', ['TRX / suspension trainer'], False),
    (r'\bRing(s)?\b|\bGymnastics? [Rr]ings?\b', ['Gymnastic rings'], False),

    # Bands
    (r'\bTheraband\b|\bResistance [Bb]and\b|\bMini [Bb]and\b|\bBand-assisted\b|\bBanded\b|\b[Bb]and\b|\bRubber band\b|\bResist band\b', ['Resistance band'], False),

    # Pull-up / hanging
    (r'\bDip [Bb]ars?\b|\bParallettes?\b', ['Dip bars'], False),
    (r'\bAb [Ss]traps?\b', ['Ab straps'], False),
    (r'\bDoor[- ]?[Mm]ount(ed)? [Pp]ull[- ]?[Uu]p [Bb]ar\b|\bDoorway [Pp]ull[- ]?[Uu]p [Bb]ar\b', ['Pull-up bar'], False),
    (r'\bPull[- ]?[Uu]p [Bb]ar\b', ['Pull-up bar'], False),

    # Plyo / boxes / bench
    # Note: "Box Jump" intentionally NOT in pattern — it's an exercise name
    # that appears in substitute text (e.g., "for lower box jump"), not equipment.
    (r'\bPlyo Box\b|\bPlyometric Box\b|\bStep[- ]?up [Bb]ox\b|\bLow [Bb]ox\b', ['Plyo box'], False),
    (r'\bMini [Hh]urdles?\b|\bAgility [Hh]urdles?\b|\bHurdles?\b', ['Mini hurdles'], False),
    (r'\bAgility [Ll]adder\b|\b[Ss]peed [Ll]adder\b', ['Agility ladder'], False),
    (r'(?<!park )(?<!Park )\bBench\b', ['Bench'], False),

    # Climbing
    (r'\bHangboard\b|\bFingerboard\b', ['Hangboard'], False),
    (r'\bCampus [Bb]oard\b', ['Campus board'], False),
    (r'\bGrip Trainer\b|\bGripper\b', ['Grip trainer'], False),
    (r'\bClimbing [Hh]olds?\b|\bSystem [Bb]oard\b|\b[Mm]oonboard\b|\b[Kk]ilter\b|\b[Tt]ension [Bb]oard\b', ['Climbing holds'], False),
    (r'\bFriction [Rr]ope\b|\b[Cc]limbing [Rr]ope\b|\bRope [Cc]limb\b', ['Climbing rope'], False),

    # Paddle
    (r'\bPaddle [Ee]rg(ometer)?\b', ['Paddle ergometer'], False),
    (r'\bPaddle\b', ['Paddle (double-blade)'], False),
    (r'\bKickboard\b', ['Kickboard'], False),

    # Carry / sled
    (r'\bSled\b|\bProwler\b|\bYoke\b', ['Weighted sled'], False),

    # Mobility
    (r'\bFoam Roller\b|\bFoam [Rr]oll\b', ['Foam roller'], False),
    (r'\bLacrosse Ball\b', ['Lacrosse ball'], False),
    (r'\bMassage Gun\b|\b[Tt]heragun\b|\b[Pp]ercussion [Mm]assager\b', ['Massage gun'], False),
    (r'\bStick Roller\b|\b[Tt]he [Ss]tick\b', ['Stick roller'], False),

    # Balance
    (r'\bBOSU\b', ['BOSU ball'], False),
    (r'\bWobble [Bb]oard\b', ['Wobble board'], False),
    (r'\bBalance [Dd]isc\b|\bAir [Dd]isc\b', ['Balance disc'], False),
    (r'\bBalance [Pp]ad\b|\bInflatable [Bb]alance|\bFoam [Pp]ad\b', ['Foam pad'], False),
    (r'\bSwiss [Bb]all\b|\bStability [Bb]all\b|\bExercise [Bb]all\b', ['Stability ball'], False),
    (r'\bSlide [Bb]oard\b|\bSlider [Bb]oard\b', ['Slider discs'], False),
    (r'\bSliding [Dd]isc\b|\b[Vv]alslides?\b|\b[Ss]liding [Hh]amstring\b|\b[Ss]liders?\b', ['Slider discs'], False),

    # Bikes
    (r'\bMountain Bike\b|\bMTB\b', ['Mountain bike'], False),
    (r'\bRoad Bike\b', ['Road bike'], False),
    (r'\bGravel Bike\b', ['Gravel bike'], False),

    # Rope / jump rope / rebounder
    (r'\bJump [Rr]ope\b|\b[Ss]kipping [Rr]ope\b', ['Jump rope'], False),
    (r'\b[Mm]ini [Tt]rampoline\b|\b[Rr]ebounder\b|\b[Rr]ebounding\b', ['Mini trampoline'], False),

    # Track (improvised — outdoor surface, not canonical equipment)
    (r'\bTrack\b(?! pant)', [], True),

    # Cross-discipline running cardio (no equipment)
    (r'\bRunning (threshold|tempo|interval|sustained)\b|\brunning (tempo|interval) equivalent\b', [], False),

    # Bodyweight indicators
    (r'\(?bodyweight\)?|\(no load\)|\bBW\b|^Bodyweight |\bunresisted\b|\bunloaded\b', [], False),
    (r'\bIncline push-?up\b|\bDeficit push-?up\b|\bdiamond push-?up\b|\bsingle-arm push-?up\b', [], False),
    (r'\bForearm (plank|side plank)\b', [], False),
    (r'\bSingle-Leg (Hip Thrust|Glute Bridge|RDL)\b|\bGoblet [Ss]quat\b|\bFarmer [Cc]arry\b', [], False),
    (r'\b(Step-back|Box) [Bb]urpee\b|\bHalf get-up\b|\bMuscle clean\b|\bCross-arm grip\b', [], False),
    (r'\bSupine figure-4\b|\bThread-the-needle\b|\bSide-lying\b|\bProne (cobra|rear delt|figure)\b', [], False),
    (r'\bHammer Curl\b', ['Dumbbell'], False),
    (r'\bWall in any room\b|\bwall in\b', [], False),
    (r'\bcountertop\b', [], True),
    (r'\bKick drill without board\b', [], False),

    # NOTE: Stairs handling is done via REAL_STAIRS_ENTRIES check before pattern matching.
    # Generic "stair" mentions that aren't real stairs → fall through to no-equipment universal.
    (r'\bStair(s|case|well)?\b|\bStep[- ]?off\b', [], False),
]

# ── Environmental patterns ─────────────────────────────────────────────────
# Backpack REMOVED — now canonical equipment.
ENVIRONMENTAL_PATTERNS = [
    r'\bgallon jugs?\b', r'\bwater jug\b', r'\bgrocery bags?\b',
    r'\b(filled )?suitcases?\b', r'\bduffel bag\b', r'\bfilled bag\b',
    r'\btree branch\b', r'\bplayground\b', r'\bmonkey bars\b',
    r'\bstairwell railing\b', r'\bhotel bed\b|\bhotel room\b',
    r'\bbed edge\b|\bbed anchor\b|\bbed frame\b|\blie .* bed\b|\bbed or (floor|carpet)\b|\bon bed\b|\bbed leg\b',
    r'\bdoor frame\b|\bdoorway\b(?! pull-up bar)|\bdoor handle\b|\bclosed door\b|\b(over|on) (top of )?door\b',
    r'\bcouch\b|\bchair\b|\bsofa\b|\bSturdy chairs?\b',
    r'^Marked floor\b|\bmarked floor\b|\bMarked tape\b',
    r'\bsand or grass\b|\bGrass hill\b|\bsand .* grass\b|\bgrassy area\b|\bopen field\b',
    r'\bpark bench\b|\bpark grass\b',
    r'\bcurb\b|\bsidewalk\b',
    r'\bhighway overpass\b|\boverpass\b',
    r'\bhallway\b|\bparking lot\b|\bparking garage\b',
    r'\bmark distance with\b|\bmark landing\b',
    r'\btowel\b', r'\bswitch hands\b',
    r'\bphysio table\b|\bmassage table\b',
    r'\bfoam mat\b|\byoga mat\b',
    r'\bany flat surface\b|\bany surface\b|\bany flat \d+m\b',
    r'\bthick book\b|\bbook on floor\b|\bbook by spine\b',
    r'\bphone or ruler\b|\btape measure\b',
    r'\bhair tie\b|\bHairband\b|\brubber bands? doubled\b|\brubber bands? around\b',
    r'\bany room\b|\bwall in\b',
    r'\bball between\b|\bsqueeze.*ball\b',
    r'\bhug a\b',
    r'\bno equipment\b|\bno resistance\b',
    r'\bbroomstick\b|\b[Pp]ole-?assisted\b',
    r'\bhammer or mallet\b',
    r'\bbrick\b|\bheavy book\b',
    r'\bbucket of (sand|rice|beans)\b|\bbag of dried\b|\bDry rice\b',
    r'\bplastic bag\b',
    r'\bpillow\b|\bknee pad from\b',
    r'\bpool noodle\b|\bnoodle\b',
    r'\binclined wooden board\b|\bWedge under\b',
    r'\bknotted bedsheet\b|\bbedsheet\b',
    r'\bMat over obstacle\b|\bfoam mat\b',
    r'\bGarmin\b|\binterval timer\b|\b[Rr]un/[Ww]alk alert\b',
    r'\bwater bottles? .* (hand|thigh)\b|\bwater bottle between\b',
    r'\bHandlebar spacer\b',
    r'\bUphill bounding\b',
    r'\bnot a true substitute\b',
    r'\bcarpet\b',
    r'\barm-only\b',
    r'\bPoles on dry land\b',
    r'\blie on floor\b|\bon floor\b',
    r'\b(reduced impact|reduced friction|lower height)\b',
    r'\bor floor\b',
    r'\banchored to (door|bed|chair|wall)\b|\banchored on (door|bed|chair|wall)\b|\banchored overhead on door\b',
    r'\bover top of door\b',
]

ENV_REGEX = re.compile('|'.join(ENVIRONMENTAL_PATTERNS), re.IGNORECASE)


# Adjunct equipment patterns — checked independently and ADD to matched_eq
# without competing with primary-equipment first-match-wins logic.
ADJUNCT_PATTERNS = [
    (r'\b[Bb]ackpack\b|\bwith pack\b|\bpack-weight\b', 'Backpack'),
    (r'\bWeight(ed)? Vest\b|\bweighted vest alternative\b', 'Weighted vest'),
]


def parse_substitute(ex_id: str, text: str) -> dict:
    text = text.strip()
    is_home = text.startswith('🏠')
    cleaned = text[1:].strip() if is_home else text
    is_env = bool(ENV_REGEX.search(cleaned))
    is_improvised = is_home or is_env

    matched_eq = []

    # ── Step 1: Adjunct equipment (Stairs, Backpack, Weighted vest) ─────
    # These are independent; they add to matched_eq without competing with
    # the primary-equipment first-match logic below.
    if is_real_stairs(ex_id, cleaned):
        matched_eq.append('Stairs')

    for pat, item in ADJUNCT_PATTERNS:
        if re.search(pat, text, re.IGNORECASE):
            if item not in matched_eq:
                matched_eq.append(item)

    # ── Step 2: Primary equipment (first match wins) ────────────────────
    # Skip patterns that match adjunct equipment (Backpack, Weighted vest)
    # since those are handled above.
    for pat, eq, override in EQUIPMENT_PATTERNS:
        if re.search(pat, text, re.IGNORECASE):
            # Skip if this pattern's equipment is purely adjunct (already handled).
            # Empty eq must NOT skip — bodyweight/absence patterns need to break the loop.
            if eq and all(item in {'Backpack', 'Weighted vest'} for item in eq):
                continue
            for item in eq:
                if item not in matched_eq:
                    matched_eq.append(item)
            if override is True:
                is_improvised = True
            break

    # Wrap into AND-OR schema
    if matched_eq:
        equipment_required = [matched_eq]
    else:
        equipment_required = []

    return {
        'substitute_text':    cleaned,
        'equipment_required': equipment_required,
        'is_improvised':      is_improvised,
    }


# ── Post-processing: Backpack → Vest disjunction, manual patches ─────────

def add_vest_alternative(data: list) -> None:
    """For groups containing Backpack, add a parallel group with Weighted vest."""
    for e in data:
        for p in e['substitutes']:
            new_groups = list(p['equipment_required'])
            for group in p['equipment_required']:
                if 'Backpack' in group:
                    vest_group = ['Weighted vest' if x == 'Backpack' else x for x in group]
                    if vest_group not in new_groups:
                        new_groups.append(vest_group)
            p['equipment_required'] = new_groups


def split_track_treadmill(data: list) -> None:
    """Split EX179 'Track or treadmill' into two separate substitute entries."""
    for e in data:
        if e['ex_id'] != 'EX179':
            continue
        new_subs = []
        for p in e['substitutes']:
            if 'Track or treadmill' in p['substitute_text']:
                # Replace with two entries
                # Track variant: outdoor, improvised
                new_subs.append({
                    'substitute_text': 'Track at goal pace',
                    'equipment_required': [],
                    'is_improvised': True,
                })
                # Treadmill variant: indoor, equipment
                new_subs.append({
                    'substitute_text': 'Treadmill at goal pace',
                    'equipment_required': [['Treadmill']],
                    'is_improvised': False,
                })
            else:
                new_subs.append(p)
        e['substitutes'] = new_subs


def patch_goblet_squat(data: list) -> None:
    """EX241 Goblet squat substitute is genuinely DB or KB. Patch the disjunction."""
    for e in data:
        if e['ex_id'] != 'EX241':
            continue
        for p in e['substitutes']:
            if 'Goblet squat (EX002 as bilateral sub)' in p['substitute_text']:
                p['equipment_required'] = [['Dumbbell'], ['Kettlebell']]


def main(xlsx_path: str = None, output_path: str = None):
    if xlsx_path is None:
        xlsx_path = Path(__file__).parent / 'AR_Exercise_Database_v17.xlsx'
    if output_path is None:
        output_path = Path(__file__).parent / 'parsed_substitutes.json'

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Exercise Master']

    parsed = []
    for r in range(3, ws.max_row + 1):
        ex_id = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        sub = ws.cell(r, 11).value
        if not ex_id or not str(ex_id).startswith('EX') or not sub:
            continue
        parts = [p.strip() for p in str(sub).split(';') if p.strip()]
        parsed.append({
            'ex_id':       ex_id,
            'name':        name,
            'substitutes': [parse_substitute(ex_id, p) for p in parts],
        })

    # Post-processing
    add_vest_alternative(parsed)
    split_track_treadmill(parsed)
    patch_goblet_squat(parsed)

    total = sum(len(e['substitutes']) for e in parsed)
    print(f"parse_substitutes v3: {len(parsed)} exercises, {total} total entries")

    with open(output_path, 'w') as f:
        json.dump(parsed, f, indent=2, ensure_ascii=False)


if __name__ == '__main__':
    import sys
    main(*sys.argv[1:])
