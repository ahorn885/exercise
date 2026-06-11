"""
clean_equipment_column.py — Pass 1 source-data cleanup for AR_Exercise_Database.

Reads v17 xlsx, applies cleanup rules to the Equipment column, writes v18 xlsx.
Produces a verbose diff log and a summary report.

Cleanup rules (locked from session decisions):
  1. Strip parenthesized content from each token (fixes bugs A/B/C)
  2. Drop tokens that violate the "canonical gear toggles only" principle
  3. Aggregate sub-component tokens into kit-level canonical entries
  4. Normalize variant names to canonical form

Idempotent: running on already-cleaned input produces identical output.
"""

import openpyxl
import re
from copy import copy
from pathlib import Path
from collections import Counter, defaultdict


# ── Cleanup rule sets ──────────────────────────────────────────────────────

# Tokens removed entirely. Either universal/situational venues, sub-components
# of canonical kits, athlete-choice clothing, or sport-skill venues we don't model.
DROP_TOKENS = frozenset({
    # Universal/situational venues
    'Floor', 'Wall', 'Marked Floor', 'Open Space', 'Outdoor', 'Indoor',
    # Sport-skill venues we don't model (only training underlying physical attributes)
    'Fencing Strip', 'Running Space', 'Shooting Range',
    # Universal water — water sports are gated by terrain (Open Water, etc.)
    'Water',
    # Universal running surfaces — running can be done anywhere
    'Track',
    # Athlete-choice clothing/gear
    'Wetsuit', 'Full Triathlon Gear Set', 'Running Shoes',
    # Sub-components of TT bike
    'Aero Bars',
    # Specialized equipment we don't model
    'Gymnastics Horse', 'Vault Box',
    # Sub-components assumed by canonical parent equipment
    'Erg Handle',
    'Canoe Seat', 'Canoe with Yoke', 'Padded Gunwales',
    'Pull-Up Bar Towel Wrap',
    'Boot Buckles',
    # Paddles assumed by boat
    'Sculling Blade', 'Two Paddles', 'Paddle', 'Paddle (double-blade)',
    'Oars', 'SUP Paddle', 'Single-Blade Paddle', 'SwimRun Paddles', 'Rowing Oar',
    # Athlete-choice swim/race kit
    'Swim Cap and Goggles', 'Spray Skirt', 'Pull Buoy',
    'Headlamp', 'Soft Flask', 'Chews', 'Cups', 'Gels', 'Knee Pad',
    # Universal AR nav gear (every AR athlete has these)
    'Compass', 'GPS', 'Topographic Map',
    # Shooting equipment (sport-skill we don't model)
    'Air Pistol', 'Laser Pistol',
    # Improvised grip training (use canonical equipment instead)
    'Rice Bucket',
    # Situational fragments (already SITUATIONAL at chunk level)
    'Visual Cue', 'Partner', 'Tandem Partner', 'Team',
    # Mental/sport-skill objects (not physical equipment)
    'Route', 'Target',
    # Too-vague tokens (flagged for manual review)
    'Machine', 'Bike',
    # Configuration tokens (athlete decides loading)
    'Bikepacking Setup',
    'Loaded Touring Bike',
    # No-equipment indicator (means absence; equivalent to empty)
    'Bodyweight',
    # Decision: cross-discipline body positions
    'Feet on Wall',
})

# Tokens flagged for manual review when dropped — log each occurrence so user
# can inspect whether the exercise still has a usable equipment requirement.
DROP_FLAGGED = frozenset({
    'Machine', 'Bike', 'Loaded Touring Bike',
})

# Tokens that aggregate into a kit-level canonical entry.
AGGREGATES = {
    # Climbing on-wall kit (rope, harness, belay device, helmet, ascender, sling)
    'Belay Device': 'Climbing gear',
    'Rope': 'Climbing gear',
    'Fixed Rope': 'Climbing gear',
    'Rope Rig': 'Climbing rope',  # exception: rope-climb-as-strength-exercise
    'Sling': 'Climbing gear',
    'Slings': 'Climbing gear',
    'Harness': 'Climbing gear',
    'Mountaineering Harness': 'Climbing gear',
    'Carabiners': 'Climbing gear',
    'Anchor Hardware': 'Climbing gear',
    'Anchor Point': 'Climbing gear',
    'Mechanical Ascender': 'Climbing gear',
    'Via Ferrata Y-Lanyard': 'Climbing gear',
    # Cross-country ski kit
    'Cross-Country Skis': 'XC ski kit',
    'Classic Cross-Country Skis': 'XC ski kit',
    'Skate Cross-Country Skis': 'XC ski kit',
    'Ski Poles on Flat Ground': 'XC ski kit',
    # Alpine touring ski kit (touring + alpine downhill = same kit functionally for AR)
    'Ski Boots': 'Touring ski kit',
    'Touring Skis': 'Touring ski kit',
    'Touring Skis with Climbing Skins': 'Touring ski kit',
    'Climbing Skins': 'Touring ski kit',
    'Ski Crampons': 'Touring ski kit',
    'Alpine Skis': 'Touring ski kit',
}

# Direct rename to existing or new canonical name.
NORMALIZATIONS = {
    'BOSU': 'BOSU ball',
    'TRX': 'TRX / suspension trainer',
    'Rack': 'Squat rack',
    'MTB Bike': 'Mountain bike',
    'MTB': 'Mountain bike',
    'Heavy Dumbbell': 'Dumbbell',
    'Light Dumbbell': 'Dumbbell',
    'Light DB': 'Dumbbell',
    'Light Plate': 'Weight plates',
    'Weight Plate': 'Weight plates',
    'Loaded Backpack': 'Backpack',
    'Loaded Packraft': 'Packraft',
    'Inclined Treadmill': 'Treadmill',
    'Inclined Treadmill at Maximum Grade': 'Treadmill',
    'Incline Bench': 'Bench',
    'Stand-Up Paddleboard': 'SUP',
    'Barbell on Thighs': 'Barbell',
    'Barbell with Plates': 'Barbell',
    # Resistance band variants (canonical: 'Resistance band', lowercase b)
    'Band': 'Resistance band',
    'Resistance Band': 'Resistance band',
    'Rubber Band': 'Resistance band',
    # Cable machine variants
    'Cable': 'Cable machine',
    'Cable Machine': 'Cable machine',
    # Bike trainer variants
    'Trainer': 'Bike trainer',
    # Weighted vest variants
    'Vest': 'Weighted vest',
    'Weight Vest': 'Weighted vest',
    # Plyo box variants
    'Box': 'Plyo box',
    # Case normalizations — source mixes capitalization for canonical names
    'Mountain Bike': 'Mountain bike',
    'Road Bike': 'Road bike',
    'Bike Trainer': 'Bike trainer',
    'Climbing Rope': 'Climbing rope',
    'Gravel Bike': 'Gravel bike',
    'Plyo Box': 'Plyo box',
}

# Tokens that split into multiple canonical tokens.
# The source xlsx uses ' or ' between tokens to indicate disjunction. Without
# AND-OR encoding for primary equipment[], we conservatively decompose to all
# named options (over-restrictive but preserves data). Where both halves drop,
# the result is empty.
SPLITS = {
    # Equipment disjunctions
    'Road Bike on Trainer': ['Road bike', 'Bike trainer'],
    'Road Bike or Trainer': ['Road bike', 'Bike trainer'],
    'Road or Gravel Bike': ['Road bike', 'Gravel bike'],
    'Road or MTB Bike': ['Road bike', 'Mountain bike'],
    'TT Bike or Road Bike on Trainer': ['TT Bike', 'Road bike', 'Bike trainer'],
    'Bench or Box': ['Bench', 'Plyo box'],
    'Floor or Bench': ['Bench'],
    # Compound aggregates (both halves map to same kit)
    'Classic or Skate Cross-Country Skis': ['XC ski kit'],
    'Cross-Country Skis or Ski Poles on Flat Ground': ['XC ski kit'],
    'Touring Skis or Alpine Skis': ['Touring ski kit'],
    'Fixed Rope or Sling': ['Climbing gear'],
    'Climb or Inclined Treadmill': ['Treadmill'],
    # Compound drops (both halves drop)
    'Canoe with Yoke or Padded Gunwales': [],
    'Fencing Strip or Marked Floor': [],
    'Fencing Strip or Open Space': [],
    'Laser Pistol or Air Pistol': [],
    'Loaded Touring Bike or Bikepacking Setup': [],
    'Partner or Visual Cue': [],
}

# Tokens that already match one of the cleanup outputs and shouldn't be flagged
# as "unchanged" garbage. Used for the diff log to know what's expected to land.
NEW_CANONICAL_REQUIRED = frozenset({
    'Climbing gear', 'XC ski kit', 'Touring ski kit', 'SUP', 'TT Bike',
})


# ── Token cleanup ──────────────────────────────────────────────────────────

PAREN_RE = re.compile(r'\s*\([^)]*\)\s*')


def strip_parens(s: str) -> str:
    """Remove all parenthesized content from a string."""
    return PAREN_RE.sub('', s).strip()


def clean_token(token: str) -> tuple[list[str], str | None]:
    """Apply cleanup rules to a single token.

    Returns:
        cleaned_tokens: list of canonical tokens (may be empty if dropped, or
                        multiple if split). Whitespace and parens already stripped.
        flag_reason:    If dropped with a manual-review flag, the reason string.
                        Otherwise None.

    Lookups are case-insensitive — handles source xlsx mixing 'Mountain Bike'
    (caps) and 'Mountain bike' (lowercase) by matching either to the canonical
    output.
    """
    # Strip parens content first — fixes bugs A/B/C and "(optional)" annotations
    t = strip_parens(token).strip()

    if not t:
        return [], None

    tl = t.lower()

    # Apply rules in priority order. Lookups via lowercased token; output
    # uses the canonical-cased value from the original rule dict.
    if tl in _SPLITS_LOWER:
        return _SPLITS_LOWER[tl], None
    if tl in _DROP_TOKENS_LOWER:
        flag = f"DROP_FLAGGED({_DROP_TOKENS_LOWER[tl]})" if tl in _DROP_FLAGGED_LOWER else None
        return [], flag
    if tl in _AGGREGATES_LOWER:
        return [_AGGREGATES_LOWER[tl]], None
    if tl in _NORMALIZATIONS_LOWER:
        return [_NORMALIZATIONS_LOWER[tl]], None

    # Token passes through unchanged
    return [t], None


# Precompute case-insensitive lookup tables. Keys are lowercased; values are
# the canonical (cased) output strings from the rule dicts.
_DROP_TOKENS_LOWER     = {t.lower(): t for t in DROP_TOKENS}
_DROP_FLAGGED_LOWER    = {t.lower(): t for t in DROP_FLAGGED}
_AGGREGATES_LOWER      = {k.lower(): v for k, v in AGGREGATES.items()}
_NORMALIZATIONS_LOWER  = {k.lower(): v for k, v in NORMALIZATIONS.items()}
_SPLITS_LOWER          = {k.lower(): v for k, v in SPLITS.items()}


def clean_equipment_string(raw: str) -> tuple[str, list[dict]]:
    """Clean an Equipment column value. Returns (cleaned_str, change_log)."""
    if not raw or not raw.strip():
        return '', []

    # Tokenize on comma. Parens content is stripped per-token in clean_token,
    # so a token like "Belay Device (ATC, GriGri)" arriving here has already
    # been split incorrectly — we handle by re-stripping parens during token cleanup.
    # But to handle the case correctly, we do paren-aware tokenization first:
    tokens = paren_aware_split(raw)

    cleaned_tokens = []
    changes = []

    for tok in tokens:
        original = tok.strip()
        if not original:
            continue

        result_tokens, flag = clean_token(original)

        # Compare normalized forms to detect changes
        original_stripped_parens = strip_parens(original).strip()

        if not result_tokens and original:
            # Dropped
            changes.append({
                'op': 'DROP',
                'from': original,
                'flag': flag,
            })
        elif result_tokens == [original_stripped_parens]:
            # Only paren stripping happened (or no change)
            if original != original_stripped_parens:
                changes.append({
                    'op': 'STRIP_PARENS',
                    'from': original,
                    'to': original_stripped_parens,
                })
            cleaned_tokens.extend(result_tokens)
        else:
            # Real transformation (normalize, aggregate, split)
            changes.append({
                'op': 'TRANSFORM',
                'from': original,
                'to': result_tokens,
            })
            cleaned_tokens.extend(result_tokens)

    # Dedupe while preserving order (aggregates may produce duplicates)
    seen = set()
    deduped = []
    for t in cleaned_tokens:
        if t not in seen:
            seen.add(t)
            deduped.append(t)

    return ', '.join(deduped), changes


def paren_aware_split(s: str) -> list[str]:
    """Split on comma, respecting parentheses. Commas inside parens don't split.

    After comma-split, strips any leading 'or ' from each token (handles
    Oxford comma constructions like 'Road Bike, MTB, or Trainer' which split
    to ['Road Bike', 'MTB', 'or Trainer']; the leading 'or ' on the last token
    is meaningless).

    Does NOT auto-split on ' or ' between tokens. The source xlsx uses ' or '
    for at least four semantically-different patterns:
      1. Independent equipment options (TT Bike or Road Bike on Trainer)
      2. Modifier disjunction (Classic or Skate Cross-Country Skis)
      3. Compound terrain names (Pool or Flat Water — already in TERRAIN_TOKENS)
      4. Compound equipment names not yet decided (Bench or Box)

    Auto-splitting was tried and breaks cases 2 and 3. Each compound is
    handled explicitly via SPLITS or NORMALIZATIONS instead.
    """
    tokens = []
    current = []
    depth = 0
    for ch in s:
        if ch == '(':
            depth += 1
            current.append(ch)
        elif ch == ')':
            depth = max(0, depth - 1)
            current.append(ch)
        elif ch == ',' and depth == 0:
            tokens.append(''.join(current).strip())
            current = []
        else:
            current.append(ch)
    if current:
        tokens.append(''.join(current).strip())

    # Strip leading "or " (Oxford comma artifact)
    tokens = [re.sub(r'^or\s+', '', t).strip() for t in tokens]

    return [t for t in tokens if t]


# ── Main ───────────────────────────────────────────────────────────────────

def main(input_path: str, output_path: str, log_path: str):
    wb = openpyxl.load_workbook(input_path)
    ws = wb['Exercise Master']

    # Find the Equipment column index
    equipment_col = None
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(2, col_idx).value
        if header and 'equipment' in str(header).strip().lower() and 'sub' not in str(header).lower():
            equipment_col = col_idx
            break

    if equipment_col is None:
        # Fallback: try column 7 (typical position)
        equipment_col = 7
        print(f"WARNING: Couldn't identify Equipment column by header; defaulting to col {equipment_col}")
    else:
        print(f"Equipment column: {equipment_col} (header: '{ws.cell(2, equipment_col).value}')")

    # Track changes
    diff_lines = []
    stats = {
        'rows_total': 0,
        'rows_changed': 0,
        'rows_unchanged': 0,
        'rows_emptied': 0,
        'tokens_before': 0,
        'tokens_after': 0,
    }
    op_counter = Counter()
    flagged_drops = defaultdict(list)  # ex_id → [token, ...]
    transformations = defaultdict(int)  # "from → to" → count

    for row_idx in range(3, ws.max_row + 1):
        ex_id = ws.cell(row_idx, 1).value
        ex_name = ws.cell(row_idx, 2).value
        if not ex_id or not str(ex_id).startswith('EX'):
            continue

        original = ws.cell(row_idx, equipment_col).value
        if not original:
            continue

        original_str = str(original).strip()
        cleaned_str, changes = clean_equipment_string(original_str)
        stats['rows_total'] += 1
        stats['tokens_before'] += len(paren_aware_split(original_str))
        stats['tokens_after'] += len(paren_aware_split(cleaned_str)) if cleaned_str else 0

        if changes:
            stats['rows_changed'] += 1
            if not cleaned_str:
                stats['rows_emptied'] += 1

            diff_lines.append(f"\n{ex_id} — {ex_name}")
            diff_lines.append(f"  BEFORE: {original_str}")
            diff_lines.append(f"  AFTER:  {cleaned_str if cleaned_str else '(empty)'}")
            for c in changes:
                op_counter[c['op']] += 1
                if c['op'] == 'DROP':
                    flag_str = f" [{c['flag']}]" if c.get('flag') else ''
                    diff_lines.append(f"    - DROP: '{c['from']}'{flag_str}")
                    if c.get('flag'):
                        flagged_drops[ex_id].append(c['from'])
                elif c['op'] == 'STRIP_PARENS':
                    diff_lines.append(f"    ~ STRIP_PARENS: '{c['from']}' → '{c['to']}'")
                elif c['op'] == 'TRANSFORM':
                    if isinstance(c['to'], list):
                        if len(c['to']) > 1:
                            diff_lines.append(f"    → SPLIT: '{c['from']}' → {c['to']}")
                            transformations[f"{c['from']} → {c['to']}"] += 1
                        else:
                            diff_lines.append(f"    → NORMALIZE: '{c['from']}' → '{c['to'][0]}'")
                            transformations[f"{c['from']} → {c['to'][0]}"] += 1
                    else:
                        diff_lines.append(f"    → TRANSFORM: '{c['from']}' → '{c['to']}'")

            # Update the cell
            ws.cell(row_idx, equipment_col).value = cleaned_str if cleaned_str else None
        else:
            stats['rows_unchanged'] += 1

    # Save the cleaned xlsx
    wb.save(output_path)

    # Write the diff log
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write("=" * 70 + "\n")
        f.write("Equipment Column Cleanup — v17 → v18\n")
        f.write("=" * 70 + "\n\n")
        f.write("## Summary\n\n")
        f.write(f"Total exercise rows:        {stats['rows_total']}\n")
        f.write(f"Rows changed:               {stats['rows_changed']}\n")
        f.write(f"Rows unchanged:             {stats['rows_unchanged']}\n")
        f.write(f"Rows emptied (no equipment after cleanup): {stats['rows_emptied']}\n")
        f.write(f"Total tokens before:        {stats['tokens_before']}\n")
        f.write(f"Total tokens after:         {stats['tokens_after']}\n")
        f.write(f"Net token reduction:        {stats['tokens_before'] - stats['tokens_after']}\n\n")

        f.write("## Operation counts\n\n")
        for op, count in op_counter.most_common():
            f.write(f"  {op}: {count}\n")

        if flagged_drops:
            f.write("\n## ⚠️  FLAGGED DROPS — manual review needed\n\n")
            f.write("These exercises had tokens dropped that may have been their primary\n")
            f.write("equipment requirement. Review whether each exercise still has a usable\n")
            f.write("Equipment column after cleanup.\n\n")
            for ex_id in sorted(flagged_drops.keys()):
                tokens = flagged_drops[ex_id]
                ex_name = next((ws.cell(r, 2).value for r in range(3, ws.max_row + 1)
                                if ws.cell(r, 1).value == ex_id), '?')
                final_eq = ws.cell(
                    next(r for r in range(3, ws.max_row + 1) if ws.cell(r, 1).value == ex_id),
                    equipment_col
                ).value or '(empty)'
                f.write(f"  {ex_id} ({ex_name}): dropped {tokens}\n")
                f.write(f"    final Equipment column: {final_eq}\n")

        f.write("\n## Top transformations\n\n")
        for transform, count in sorted(transformations.items(), key=lambda x: -x[1])[:20]:
            f.write(f"  {count:3} × {transform}\n")

        f.write("\n## Per-exercise diff\n")
        f.write("\n".join(diff_lines))

    print(f"\nWrote cleaned xlsx: {output_path}")
    print(f"Wrote diff log:     {log_path}")
    print()
    print(f"Summary: {stats['rows_changed']}/{stats['rows_total']} rows changed; "
          f"{stats['tokens_before']}→{stats['tokens_after']} tokens "
          f"({stats['tokens_before'] - stats['tokens_after']} dropped)")
    if flagged_drops:
        print(f"  ⚠️  {len(flagged_drops)} exercises with FLAGGED drops — review log")
    if stats['rows_emptied']:
        print(f"  ⚠️  {stats['rows_emptied']} exercises now have empty Equipment columns — review")


if __name__ == '__main__':
    import sys
    args = sys.argv[1:]
    main(
        args[0] if args else 'AR_Exercise_Database_v17.xlsx',
        args[1] if len(args) > 1 else 'AR_Exercise_Database_v18.xlsx',
        args[2] if len(args) > 2 else 'cleanup_diff_log.txt',
    )
