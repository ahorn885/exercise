"""
Pass 2 cleanup: v18 → v19
- Applies all Andy's exercise-level corrections
- Resolves all 7 unresolved Pass 2 vocab tokens
- Cascades deletions to Sport-Exercise Map
"""
import pandas as pd
from openpyxl import load_workbook
import copy

SRC = "/mnt/project/AR_Exercise_Database_v18.xlsx"
DST = "/mnt/user-data/outputs/AR_Exercise_Database_v19.xlsx"

# ── 1. DELETIONS ──────────────────────────────────────────────────────────────
DELETE_IDS = {
    'EX059',  # Neck Isometric Hold — drop per Andy
    # Technique drills Andy wants dropped
    'EX132','EX133','EX134','EX135','EX136','EX137',
    'EX141','EX143','EX145','EX146','EX147','EX151',
    'EX161','EX177','EX181','EX182','EX187','EX188',
    'EX189','EX190','EX191','EX192','EX193','EX198',
    'EX202','EX204','EX205','EX206','EX207','EX208',
    'EX209','EX210','EX211',
}

# ── 2. EQUIPMENT OVERRIDES (post-deletion) ────────────────────────────────────
EQUIP_OVERRIDES = {
    # Remove wrongly-mapped anchor/climbing gear
    'EX020': '',   # Nordic Hamstring Curl — no equipment needed
    'EX064': '',   # Reverse Nordic Curl — no equipment needed
    # BW + DB/KB options
    'EX022': 'Dumbbell, Kettlebell',   # Reverse Lunge (DB or BW)
    'EX023': 'Dumbbell, Kettlebell',   # Lateral Lunge (DB or BW)
    'EX025': 'Dumbbell, Kettlebell',   # Single-Leg Calf Raise (Loaded)
    'EX038': 'Dumbbell, Kettlebell',   # Split Squat ISO Hold
    # Drop venue/prop tokens
    'EX056': '',   # Reactive Direction Change — drop Cones
    # Bike cardio: all bike types valid
    'EX073': 'Road bike, Mountain bike, Bike trainer, TT Bike, Gravel bike',
    'EX074': 'Road bike, Mountain bike, Bike trainer, TT Bike, Gravel bike',
    'EX075': 'Road bike, Mountain bike, Bike trainer, TT Bike, Gravel bike',
    # Grip training — add the missing equipment
    'EX104': 'Rice bucket',
    # Climbing wall as venue toggle
    'EX114': 'Climbing Wall',
    # Add KB option
    'EX117': 'Plyo box, Dumbbell, Kettlebell, Weighted vest',
    'EX119': 'Dumbbell, Barbell, Kettlebell, Weighted vest',
    # Swim — add Pull buoy
    'EX126': 'Pool, Pull buoy',
    # Add Weighted vest to BW exercises
    'EX216': 'Weighted vest',   # Plank (Front)
    'EX219': 'Weighted vest',   # Side Plank
    'EX228': 'Weighted vest',   # Push-Up (Bodyweight)
    'EX238': 'Weighted vest',   # Burpee
    # Add KB to Lateral Raise
    'EX233': 'Dumbbell, Kettlebell, Resistance band',
}

# ── 3. PASS 2 TOKEN NORMALIZATIONS ───────────────────────────────────────────
# Applied as token-level replacements in Equipment column (surviving rows only)
TOKEN_RULES = [
    # Trekking context: bare "Poles" → "Trekking Poles"
    # Ski context: "Poles" adjacent to ski kit → drop (ski poles are part of ski kit)
    # Mountaineering tokens → aggregate
    ('Crampons',            'Mountaineering kit'),
    ('Ice Axe',             'Mountaineering kit'),
    ('Mountaineering Boots','Mountaineering kit'),
    # Packraft
    ('Inflatable Raft',     'Packraft'),
    # Venue noise in raft exercises
    ('Whitewater',          ''),   # drop
    ('Snow Slope',          ''),   # drop
]

def apply_token_rules(equip_str, ex_id):
    if not isinstance(equip_str, str) or equip_str.strip() == '':
        return equip_str
    tokens = [t.strip() for t in equip_str.split(',')]
    out = []
    for tok in tokens:
        replaced = tok
        for src, dst in TOKEN_RULES:
            if replaced == src:
                replaced = dst
                break
        # Poles: ski context (touring ski kit present in original string) → drop
        if replaced == 'Poles':
            if 'Touring ski kit' in equip_str or 'XC ski kit' in equip_str:
                replaced = ''  # ski poles fold into ski kit
            else:
                replaced = 'Trekking Poles'
        if replaced:
            out.append(replaced)
    # De-dup while preserving order, consolidate duplicate Mountaineering kit
    seen = set()
    deduped = []
    for tok in out:
        if tok not in seen:
            seen.add(tok)
            deduped.append(tok)
    return ', '.join(deduped)

# ── 4. EXECUTE ────────────────────────────────────────────────────────────────
wb = load_workbook(SRC)

# ─ Exercise Master ────────────────────────────────────────────────────────────
ws_em = wb['Exercise Master']

# Find header row (row 2 in openpyxl = index 2) and column indices
header_row = 2
cols = {ws_em.cell(row=header_row, column=c).value: c for c in range(1, ws_em.max_column+1)}
id_col   = cols['Exercise ID']
eq_col   = cols['Equipment']

rows_to_delete = []
for row in ws_em.iter_rows(min_row=header_row+1, max_row=ws_em.max_row):
    ex_id = row[id_col-1].value
    if ex_id in DELETE_IDS:
        rows_to_delete.append(row[0].row)

# Delete from bottom to preserve row numbers
for r in sorted(rows_to_delete, reverse=True):
    ws_em.delete_rows(r)

# Now apply equipment overrides and token normalizations
for row in ws_em.iter_rows(min_row=header_row+1, max_row=ws_em.max_row):
    ex_id = row[id_col-1].value
    if ex_id is None:
        continue
    eq_cell = row[eq_col-1]
    current = eq_cell.value if eq_cell.value else ''

    if ex_id in EQUIP_OVERRIDES:
        new_val = EQUIP_OVERRIDES[ex_id] or None
    else:
        new_val = apply_token_rules(current, ex_id) or None

    eq_cell.value = new_val

# ─ Sport-Exercise Map ─────────────────────────────────────────────────────────
ws_sem = wb['Sport-Exercise Map']
sem_header = 2
sem_cols = {ws_sem.cell(row=sem_header, column=c).value: c for c in range(1, ws_sem.max_column+1)}
sem_id_col = sem_cols['Exercise ID']

sem_rows_to_delete = []
for row in ws_sem.iter_rows(min_row=sem_header+1, max_row=ws_sem.max_row):
    ex_id = row[sem_id_col-1].value
    if ex_id in DELETE_IDS:
        sem_rows_to_delete.append(row[0].row)

for r in sorted(sem_rows_to_delete, reverse=True):
    ws_sem.delete_rows(r)

wb.save(DST)
print(f"Saved {DST}")
print(f"Exercise Master deletions: {len(rows_to_delete)}")
print(f"Sport-Exercise Map deletions: {len(sem_rows_to_delete)}")
